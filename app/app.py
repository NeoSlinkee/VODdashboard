"""
Support Engineer Report Tracking System v2.0
Automated data ingestion from Ruckus + Freshdesk
Local web application with health dashboard

Run with: python app.py
Access at: http://localhost:5000
"""

from functools import wraps
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, Response, send_from_directory, abort
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
import os
import csv
import json
import re
import requests
import smtplib
from email.message import EmailMessage
from difflib import get_close_matches
from io import StringIO, BytesIO
from openpyxl import Workbook
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'vod-support-tracker-2026-dev')

# --- Database URL handling (Neon Postgres or local SQLite) ---
_db_url = os.environ.get('DATABASE_URL', 'sqlite:///reports.db')
# Normalise to psycopg3 dialect (postgresql+psycopg://) so the driver works on
# any Python version including 3.14.  psycopg2 C extension is incompatible with 3.14.
if _db_url.startswith('postgres://'):
    _db_url = _db_url.replace('postgres://', 'postgresql+psycopg://', 1)
elif _db_url.startswith('postgresql://') and '+' not in _db_url.split('://')[0]:
    _db_url = _db_url.replace('postgresql://', 'postgresql+psycopg://', 1)
app.config['SQLALCHEMY_DATABASE_URI'] = _db_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Engine options: connection pooling + SSL for Neon Postgres
if _db_url.startswith('postgresql'):
    app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
        'pool_pre_ping': True,      # detect stale connections
        'pool_size': 5,
        'pool_recycle': 300,        # recycle connections every 5 min
        'max_overflow': 2,
        'connect_args': {
            'sslmode': 'require',
        },
    }

app.config['REFERENCE_DOCS_FOLDER'] = os.path.join(app.instance_path, 'reference_docs')

# Session expires after 8 hours of inactivity
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=8)

# Shared secret for cron/scheduled job endpoints
CRON_SECRET = os.environ.get('CRON_SECRET', '')

# Base URL for this deployment — used in email links and cron callbacks
APP_BASE_URL = os.environ.get('APP_BASE_URL', 'http://localhost:5000').rstrip('/')

db = SQLAlchemy(app)

login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Please log in to access this page.'
login_manager.login_message_category = 'info'

# ============== Assigned Sites ==============
# Sites assigned to this support engineer for monitoring
# Zone names should match exactly as they appear in Ruckus controller

ASSIGNED_SITES = {
    'within_reach': [
        # Sandton
        {'name': 'Garden Court Morningside Sandton', 'ruckus_zones': ['Garden Court Morningside Sandton', 'GC Morningside']},
        {'name': 'InterContinental Sandton', 'ruckus_zones': ['InterConti Sandton', 'InterConti Sandton - Staff', 'IC Sandton']},
        # Pretoria
        {'name': 'Riverside Sun', 'ruckus_zones': ['Riverside Sun']},
        {'name': 'Stay Easy Pretoria', 'ruckus_zones': ['Stay Easy Pretoria', 'StayEasy Pretoria']},
        {'name': 'Southern Sun Pretoria', 'ruckus_zones': ['Southern Sun Pretoria', 'SS Pretoria']},
        # East Gate
        {'name': 'Garden Court EastGate', 'ruckus_zones': ['Garden Court EastGate', 'GC EastGate']},
        {'name': 'Stay Easy EastGate', 'ruckus_zones': ['Stay Easy EastGate', 'StayEasy EastGate']},
        # OR Tambo
        {'name': 'Garden Court OR Tambo', 'ruckus_zones': ['Garden Court OR Tambo', 'GC OR Tambo']},
        {'name': 'InterContinental OR Tambo', 'ruckus_zones': ['InterConti OR Tambo', 'IC OR Tambo', 'IC Airport']},
        {'name': 'Southern Sun OR Tambo', 'ruckus_zones': ['Southern Sun OR Tambo', 'SS OR Tambo']},
    ],
    'far_reach': [
        # Durban
        {'name': 'Stay Easy Pietermaritzburg', 'ruckus_zones': ['Stay Easy Pietermaritzburg', 'StayEasy PMB']},
        {'name': 'Garden Court South Beach', 'ruckus_zones': ['Garden Court South Beach', 'GC South Beach']},
        {'name': 'The Edward Durban', 'ruckus_zones': ['The Edward', 'Edward Durban']},
        {'name': 'Garden Court Marine Parade', 'ruckus_zones': ['Garden Court Marine Parade', 'GC Marine Parade']},
        {'name': 'SS Elangeni & Maharani', 'ruckus_zones': ['SS Elangeni', 'Elangeni', 'Maharani', 'SS Maharani']},
        {'name': 'Suncoast Hotel and Towers', 'ruckus_zones': ['Suncoast Hotel', 'Suncoast Towers', 'Suncoast']},
        {'name': 'The Ridge Hotel', 'ruckus_zones': ['The Ridge', 'Ridge Hotel']},
    ]
}

def normalize_site_name(name):
    """Normalize site/zone names to improve matching across naming variants."""
    if not name:
        return ''
    normalized = name.lower().strip()
    normalized = normalized.replace('&', ' and ')
    normalized = normalized.replace('-', ' ')
    normalized = normalized.replace('_', ' ')
    normalized = re.sub(r'[()\[\]{}]', ' ', normalized)
    normalized = re.sub(r'[^a-z0-9\s]', ' ', normalized)
    normalized = re.sub(r'\s+', ' ', normalized).strip()
    return normalized


# Manual alias overrides for known naming gaps across controllers/exports.
SITE_NAME_MAP = {
    'garden court morningside': 'Garden Court Morningside Sandton',
    'garden court morningside staff': 'Garden Court Morningside Sandton',
    'gc morningside': 'Garden Court Morningside Sandton',
    'gc or tambo staff': 'Garden Court OR Tambo',
    'garden court or tambo staff': 'Garden Court OR Tambo',
    'interconti or tambo': 'InterContinental OR Tambo',
    'ic or tambo': 'InterContinental OR Tambo',
    'southern sun elangeni': 'SS Elangeni & Maharani',
    'southern sun elangeni staff': 'SS Elangeni & Maharani',
    'ss elangeni': 'SS Elangeni & Maharani',
    'ss maharani': 'SS Elangeni & Maharani',
    'gc marine parade': 'Garden Court Marine Parade',
    'gc south beach': 'Garden Court South Beach',
    'se pretoria': 'Stay Easy Pretoria',
    'se eastgate': 'Stay Easy EastGate',
    'ss pretoria': 'Southern Sun Pretoria',
    'ss or tambo': 'Southern Sun OR Tambo',
}

# APs with these keywords are treated as maintenance/decommissioned and excluded
# from health scoring totals.
MAINTENANCE_AP_KEYWORDS = ('faulty', 'replaced', 'spare', 'cable fault', 'decommissioned')

DEFAULT_PUBLIC_HOLIDAYS = {
    '2026-01-01': 'New Year\'s Day',
    '2026-03-21': 'Human Rights Day',
    '2026-04-03': 'Good Friday',
    '2026-04-06': 'Family Day',
    '2026-04-27': 'Freedom Day',
    '2026-05-01': 'Workers\' Day',
    '2026-06-16': 'Youth Day',
    '2026-08-09': 'National Women\'s Day',
    '2026-09-24': 'Heritage Day',
    '2026-12-16': 'Day of Reconciliation',
    '2026-12-25': 'Christmas Day',
    '2026-12-26': 'Day of Goodwill'
}

ESCALATION_TARGETS = ['Dexter', 'Armand']


# Build matching indexes for exact, normalized, and fuzzy site matching.
SITE_BY_CANONICAL = {}
ALIAS_TO_SITE = {}

for region, sites in ASSIGNED_SITES.items():
    for site in sites:
        canonical = site['name']
        canonical_norm = normalize_site_name(canonical)
        site_info = {'name': canonical, 'region': region}

        SITE_BY_CANONICAL[canonical] = site_info
        ALIAS_TO_SITE[canonical_norm] = site_info

        for zone in site['ruckus_zones']:
            ALIAS_TO_SITE[normalize_site_name(zone)] = site_info

for alias, canonical in SITE_NAME_MAP.items():
    if canonical in SITE_BY_CANONICAL:
        ALIAS_TO_SITE[normalize_site_name(alias)] = SITE_BY_CANONICAL[canonical]

# ============== Configuration ==============

class Config(db.Model):
    """System configuration storage."""
    id = db.Column(db.Integer, primary_key=True)
    key = db.Column(db.String(100), unique=True, nullable=False)
    value = db.Column(db.Text)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow)

def get_config(key, default=None):
    config = Config.query.filter_by(key=key).first()
    return config.value if config else default


class User(UserMixin, db.Model):
    """Application login users."""
    __tablename__ = 'app_user'  # 'user' is reserved in PostgreSQL
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(256), nullable=False)
    role = db.Column(db.String(20), default='viewer')
    active = db.Column(db.Boolean, default=True)
    agent_id = db.Column(db.Integer, db.ForeignKey('agent.id'), nullable=True)
    created_by = db.Column(db.Integer, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    last_login = db.Column(db.DateTime, nullable=True)

    agent = db.relationship('Agent', backref='users')

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

    def get_id(self):
        return str(self.id)

    @property
    def is_admin(self):
        return self.role in ('admin', 'superadmin')

    @property
    def is_superadmin(self):
        return self.role == 'superadmin'

    def can_manage(self, other_user):
        """Return True if this user can modify/delete the other user."""
        if self.id == other_user.id:
            return False
        if self.is_superadmin:
            return True
        if self.role == 'admin' and other_user.role == 'viewer':
            return True
        return False


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


def get_active_agent_id():
    """Return the current user's agent_id, or None if admin (unscoped)."""
    if current_user.is_authenticated and not current_user.is_admin:
        return current_user.agent_id
    return None


def get_active_agent_name():
    """Return the current user's agent name, or None if admin."""
    if current_user.is_authenticated and not current_user.is_admin and current_user.agent:
        return current_user.agent.name
    return None


def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin:
            abort(403)
        return f(*args, **kwargs)
    return decorated_function


@app.context_processor
def inject_brand():
    """Inject branding variables into every template context."""
    return {
        'APP_NAME': get_config('app_name', 'VOD Operations Portal'),
        'APP_LOGO': get_config('app_logo'),       # base64 data-URI or None
        'COMPANY_NAME': get_config('company_name', 'Vodacom'),
    }


@app.errorhandler(403)
def forbidden(e):
    return render_template('login.html', forbidden=True), 403


@app.before_request
def require_login():
    open_endpoints = {'login', 'logout', 'static', 'api_health', 'api_import_ruckus',
                       'cron_freshdesk_sync', 'cron_ruckus_import', 'cron_health'}
    if request.endpoint not in open_endpoints and not current_user.is_authenticated:
        return redirect(url_for('login', next=request.path))

def set_config(key, value):
    config = Config.query.filter_by(key=key).first()
    if config:
        config.value = value
        config.updated_at = datetime.utcnow()
    else:
        config = Config(key=key, value=value)
        db.session.add(config)
    db.session.commit()

# ============== Database Models ==============

class Agent(db.Model):
    """Support engineers."""
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(120))
    freshdesk_agent_id = db.Column(db.String(50))  # For API integration
    active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    # Standby / roster fields — replaces the old hardcoded STANDBY_ENGINEERS dict
    roster_enabled = db.Column(db.Boolean, default=True)   # eligible for duty roster
    standby_color = db.Column(db.String(20), default='#2a2f33')  # calendar cell bg colour
    standby_text_color = db.Column(db.String(20), default='#ffffff')
    standby_label = db.Column(db.String(20), default='secondary')  # Bootstrap label class
    standby_start_date = db.Column(db.Date, nullable=True)  # when they joined the roster
    standby_end_date = db.Column(db.Date, nullable=True)    # None = still active
    standby_responsibilities = db.Column(db.String(300))

class Site(db.Model):
    """Assigned sites to monitor."""
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False, unique=True)
    region = db.Column(db.String(50))  # 'within_reach' or 'far_reach'
    ruckus_zone_name = db.Column(db.String(200))  # Name as it appears in Ruckus
    total_aps = db.Column(db.Integer, default=0)
    active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    monitoring_logs = db.relationship('MonitoringLog', backref='site', lazy=True)

class MonitoringLog(db.Model):
    """Auto-populated from Ruckus data."""
    id = db.Column(db.Integer, primary_key=True)
    site_id = db.Column(db.Integer, db.ForeignKey('site.id'), nullable=False)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)
    
    total_aps = db.Column(db.Integer, default=0)
    aps_online = db.Column(db.Integer, default=0)
    aps_offline = db.Column(db.Integer, default=0)
    
    status = db.Column(db.String(20))  # healthy, warning, critical
    alerts = db.Column(db.Text)  # Any alerts from Ruckus
    
    auto_imported = db.Column(db.Boolean, default=True)
    notes = db.Column(db.Text)  # Manual notes only


class MonitoringEscalation(db.Model):
    """Escalation tracking for monitoring issues."""
    id = db.Column(db.Integer, primary_key=True)
    monitoring_log_id = db.Column(db.Integer, db.ForeignKey('monitoring_log.id'), nullable=False)
    escalated = db.Column(db.Boolean, default=True)
    escalated_to = db.Column(db.String(100), nullable=False)
    escalation_reason = db.Column(db.Text)
    telegram_sent = db.Column(db.Boolean, default=False)
    escalated_at = db.Column(db.DateTime, default=datetime.utcnow)
    resolved_at = db.Column(db.DateTime)
    resolution_notes = db.Column(db.Text)

    monitoring_log = db.relationship('MonitoringLog', backref='escalations')

class SiteVisit(db.Model):
    """Physical site visits (manual entry)."""
    id = db.Column(db.Integer, primary_key=True)
    agent_id = db.Column(db.Integer, db.ForeignKey('agent.id'), nullable=False)
    site_id = db.Column(db.Integer, db.ForeignKey('site.id'))
    site_name = db.Column(db.String(200))  # Fallback if site not in DB
    location = db.Column(db.String(200))
    date = db.Column(db.Date, nullable=False)
    discussion_topics = db.Column(db.Text)
    info_obtained = db.Column(db.Text)
    follow_up_actions = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    agent = db.relationship('Agent', backref='site_visits')
    site = db.relationship('Site', backref='visits')


class ServiceCall(db.Model):
    """Individual service call / ticket interaction (manual entry)."""
    id = db.Column(db.Integer, primary_key=True)
    agent_id = db.Column(db.Integer, db.ForeignKey('agent.id'), nullable=False)
    date = db.Column(db.Date, nullable=False)
    ticket_number = db.Column(db.String(50))
    summary = db.Column(db.String(500))
    status = db.Column(db.String(20), default='open')  # open, closed, escalated
    resolution = db.Column(db.Text)
    time_spent = db.Column(db.Integer)  # minutes
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    agent = db.relationship('Agent', backref='service_calls')

class TicketStats(db.Model):
    """Daily ticket statistics (auto from Freshdesk or manual)."""
    id = db.Column(db.Integer, primary_key=True)
    agent_id = db.Column(db.Integer, db.ForeignKey('agent.id'), nullable=False)
    date = db.Column(db.Date, nullable=False)
    
    tickets_handled = db.Column(db.Integer, default=0)
    tickets_closed = db.Column(db.Integer, default=0)
    tickets_escalated = db.Column(db.Integer, default=0)
    tickets_pending = db.Column(db.Integer, default=None)  # live open+pending count snapshot
    
    auto_imported = db.Column(db.Boolean, default=False)
    source = db.Column(db.String(50))  # 'freshdesk', 'manual'
    
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    agent = db.relationship('Agent', backref='ticket_stats')

class WeeklyReport(db.Model):
    """Weekly summary reports."""
    id = db.Column(db.Integer, primary_key=True)
    agent_id = db.Column(db.Integer, db.ForeignKey('agent.id'), nullable=False)
    week_ending = db.Column(db.Date, nullable=False)
    
    # Auto-calculated
    sites_visited = db.Column(db.Integer, default=0)
    sites_monitored = db.Column(db.Integer, default=0)
    aps_offline_total = db.Column(db.Integer, default=0)
    critical_incidents = db.Column(db.Integer, default=0)
    tickets_handled = db.Column(db.Integer, default=0)
    tickets_closed = db.Column(db.Integer, default=0)
    
    # Manual additions
    achievements = db.Column(db.Text)
    challenges = db.Column(db.Text)
    notes = db.Column(db.Text)
    
    auto_generated = db.Column(db.Boolean, default=True)
    submitted_at = db.Column(db.DateTime)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    agent = db.relationship('Agent', backref='weekly_reports')

class ImportLog(db.Model):
    """Track data imports."""
    id = db.Column(db.Integer, primary_key=True)
    source = db.Column(db.String(50))  # 'ruckus', 'freshdesk'
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)
    records_processed = db.Column(db.Integer, default=0)
    records_matched = db.Column(db.Integer, default=0)
    status = db.Column(db.String(20))  # 'success', 'partial', 'failed'
    details = db.Column(db.Text)


def _build_standby_engineers_map():
    """Build the STANDBY_ENGINEERS-style dict dynamically from the Agent table.

    Returns a dict keyed by agent name — same shape the roster/standby templates
    expect — so existing template code works without changes.
    """
    try:
        agents = Agent.query.filter_by(active=True, roster_enabled=True).all()
    except Exception:
        return {}
    result = {}
    for a in agents:
        result[a.name] = {
            'color': a.standby_color or '#2a2f33',
            'text':  a.standby_text_color or '#ffffff',
            'label': a.standby_label or 'secondary',
            'start_date': a.standby_start_date.isoformat() if a.standby_start_date else '2025-01-01',
            'end_date':   a.standby_end_date.isoformat()   if a.standby_end_date   else None,
            'responsibilities': a.standby_responsibilities or '',
        }
    return result


def get_standby_engineers():
    """Return the live standby engineers map (DB-driven, no hardcoding)."""
    return _build_standby_engineers_map()


class DutyRoster(db.Model):
    """Monthly night-standby roster — one engineer on call per day."""
    id = db.Column(db.Integer, primary_key=True)
    date = db.Column(db.Date, nullable=False, unique=True)
    agent_name = db.Column(db.String(100), nullable=False)  # Chris / Kudzayi / Koketso
    notes = db.Column(db.String(500))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


class StandbyClaim(db.Model):
    """Individual standby call-out entries for a claim form submission."""
    id = db.Column(db.Integer, primary_key=True)
    agent_name = db.Column(db.String(100), nullable=False)
    date = db.Column(db.Date, nullable=False)
    call_received = db.Column(db.String(10))   # HH:MM string
    start_time = db.Column(db.String(10), nullable=False)   # HH:MM string
    end_time = db.Column(db.String(10), nullable=False)     # HH:MM string
    hours = db.Column(db.Float, default=0.0)   # calculated on save
    description = db.Column(db.String(500))
    ticket_reference = db.Column(db.String(100))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


class LeaveRequest(db.Model):
    """Leave requests with approval workflow and standby coverage updates."""
    id = db.Column(db.Integer, primary_key=True)
    agent_name = db.Column(db.String(100), nullable=False)
    start_date = db.Column(db.Date, nullable=False)
    end_date = db.Column(db.Date, nullable=False)
    reason = db.Column(db.Text)
    status = db.Column(db.String(20), default='pending')  # pending/approved/rejected
    approver = db.Column(db.String(100), default='Luke')
    coverage_agent = db.Column(db.String(100))
    decision_notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    decided_at = db.Column(db.DateTime)


class ReferenceDocument(db.Model):
    """Stored reference files for quick lookup (e.g. monday/controller access docs)."""
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    source = db.Column(db.String(50), default='manual')
    notes = db.Column(db.Text)
    original_filename = db.Column(db.String(255), nullable=False)
    stored_filename = db.Column(db.String(255), nullable=False, unique=True)
    uploaded_at = db.Column(db.DateTime, default=datetime.utcnow)

# ============== Helper Functions ==============

def get_week_ending(date=None):
    if date is None:
        date = datetime.now()
    days_until_friday = (4 - date.weekday()) % 7
    if days_until_friday == 0 and date.weekday() == 4:
        return date.date()
    return (date + timedelta(days=days_until_friday)).date()

def get_week_start(week_ending):
    return week_ending - timedelta(days=4)

def calculate_status(total_aps, aps_offline):
    """Calculate health status based on AP offline percentage."""
    if total_aps == 0:
        return 'unknown'
    offline_pct = (aps_offline / total_aps) * 100
    if aps_offline == 0:
        return 'healthy'
    elif offline_pct < 20:
        return 'warning'
    else:
        return 'critical'


def parse_alert_metadata(alerts_text):
    """Parse structured metadata from MonitoringLog.alerts."""
    default = {
        'confidence': 'low',
        'source_basis': 'unknown',
        'data_scope': 'full_snapshot',
        'excluded_aps': 0,
        'excluded_examples': [],
        'offline_ap_count': 0,
        'offline_ap_examples': [],
        'new_offline_count': 0,
        'new_offline_examples': [],
        'restored_count': 0,
        'status_note': ''
    }
    if not alerts_text:
        return default
    try:
        parsed = json.loads(alerts_text)
        if isinstance(parsed, dict):
            return {
                'confidence': parsed.get('confidence', 'low'),
                'source_basis': parsed.get('source_basis', 'unknown'),
                'data_scope': parsed.get('data_scope', 'full_snapshot'),
                'excluded_aps': int(parsed.get('excluded_aps', 0) or 0),
                'excluded_examples': parsed.get('excluded_examples', []) or [],
                'offline_ap_count': int(parsed.get('offline_ap_count', 0) or 0),
                'offline_ap_examples': parsed.get('offline_ap_examples', []) or [],
                'new_offline_count': int(parsed.get('new_offline_count', 0) or 0),
                'new_offline_examples': parsed.get('new_offline_examples', []) or [],
                'restored_count': int(parsed.get('restored_count', 0) or 0),
                'status_note': parsed.get('status_note', '') or ''
            }
    except Exception:
        pass
    return default


def _ensure_reference_docs_folder():
    os.makedirs(app.config['REFERENCE_DOCS_FOLDER'], exist_ok=True)


def _is_allowed_reference_file(filename):
    allowed_extensions = {'xlsx', 'xls', 'csv', 'pdf', 'txt', 'md', 'json'}
    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
    return ext in allowed_extensions


def _load_reference_table_preview(doc, max_rows=250, max_cols=30):
    """Return a table preview for csv/xlsx/xls files."""
    file_path = os.path.join(app.config['REFERENCE_DOCS_FOLDER'], doc.stored_filename)
    if not os.path.exists(file_path):
        return {'error': 'Stored file could not be found.'}

    ext = doc.stored_filename.rsplit('.', 1)[-1].lower() if '.' in doc.stored_filename else ''
    rows = []

    try:
        if ext in {'csv', 'txt'}:
            try:
                with open(file_path, 'r', encoding='utf-8-sig', newline='') as f:
                    sample = f.read(2048)
                    f.seek(0)
                    try:
                        dialect = csv.Sniffer().sniff(sample)
                    except Exception:
                        dialect = csv.excel
                    reader = csv.reader(f, dialect=dialect)
                    for i, row in enumerate(reader):
                        if i >= max_rows:
                            break
                        rows.append([str(cell) if cell is not None else '' for cell in row[:max_cols]])
            except UnicodeDecodeError:
                with open(file_path, 'r', encoding='latin-1', newline='') as f:
                    reader = csv.reader(f)
                    for i, row in enumerate(reader):
                        if i >= max_rows:
                            break
                        rows.append([str(cell) if cell is not None else '' for cell in row[:max_cols]])
        elif ext in {'xlsx', 'xls'}:
            try:
                load_workbook = __import__('openpyxl').load_workbook
            except Exception:
                return {'error': 'Spreadsheet preview requires openpyxl. Please install it or upload as CSV.'}

            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= max_rows:
                    break
                rows.append([str(cell) if cell is not None else '' for cell in list(row)[:max_cols]])
            wb.close()
        else:
            return {'error': f'Preview is currently available for CSV/XLSX files. This file is .{ext or "unknown"}.'}
    except Exception as exc:
        return {'error': f'Could not read file: {exc}'}

    if not rows:
        return {'headers': [], 'rows': [], 'truncated': False, 'row_count': 0}

    max_width = max(len(r) for r in rows)
    normalized_rows = [r + [''] * (max_width - len(r)) for r in rows]
    headers = normalized_rows[0]
    data_rows = normalized_rows[1:]
    truncated = len(normalized_rows) >= max_rows

    return {
        'headers': headers,
        'rows': data_rows,
        'truncated': truncated,
        'row_count': len(normalized_rows)
    }


def calculate_weekly_monitoring_metrics(week_start_dt, week_end_dt):
    """Compute deduplicated weekly monitoring totals for reporting."""
    logs = MonitoringLog.query.filter(
        MonitoringLog.timestamp >= week_start_dt,
        MonitoringLog.timestamp <= week_end_dt
    ).all()

    sites_monitored = len({log.site_id for log in logs if log.site_id is not None})

    # AP offline dedupe: prefer unique AP identifiers from metadata, otherwise
    # fallback to max offline count per site for the week.
    unique_offline_aps = set()
    sites_with_examples = set()
    fallback_site_counts = {}

    for log in logs:
        if log.site_id is None:
            continue

        fallback_site_counts[log.site_id] = max(
            fallback_site_counts.get(log.site_id, 0),
            int(log.aps_offline or 0)
        )

        meta = parse_alert_metadata(log.alerts)
        examples = [str(ap).strip() for ap in (meta.get('offline_ap_examples') or []) if str(ap).strip()]
        if examples:
            sites_with_examples.add(log.site_id)
            for ap_name in examples:
                unique_offline_aps.add(f"{log.site_id}:{ap_name.lower()}")

    aps_offline_total = len(unique_offline_aps)
    for site_id, count in fallback_site_counts.items():
        if site_id not in sites_with_examples:
            aps_offline_total += count

    # Critical incidents dedupe: one incident per site per day when critical.
    critical_incidents = len({
        (log.site_id, log.timestamp.date())
        for log in logs
        if log.site_id is not None and log.status == 'critical'
    })

    return {
        'sites_monitored': sites_monitored,
        'aps_offline_total': aps_offline_total,
        'critical_incidents': critical_incidents
    }

def get_site_health_summary():
    """Get current health status of all sites."""
    sites = Site.query.filter_by(active=True).all()
    summary = {'healthy': 0, 'warning': 0, 'critical': 0, 'unknown': 0}
    site_status = []
    
    for site in sites:
        # Get latest monitoring log
        latest = MonitoringLog.query.filter_by(site_id=site.id)\
            .order_by(MonitoringLog.timestamp.desc()).first()
        
        if latest:
            status = latest.status or calculate_status(latest.total_aps, latest.aps_offline)
            meta = parse_alert_metadata(latest.alerts)
            site_status.append({
                'site': site,
                'status': status,
                'total_aps': latest.total_aps,
                'aps_offline': latest.aps_offline,
                'last_checked': latest.timestamp,
                'confidence': meta['confidence'],
                'source_basis': meta['source_basis'],
                'excluded_aps': meta['excluded_aps'],
                'excluded_examples': meta['excluded_examples']
            })
            summary[status] = summary.get(status, 0) + 1
        else:
            site_status.append({
                'site': site,
                'status': 'unknown',
                'total_aps': site.total_aps,
                'aps_offline': 0,
                'last_checked': None,
                'confidence': 'low',
                'source_basis': 'no_data',
                'excluded_aps': 0,
                'excluded_examples': []
            })
            summary['unknown'] += 1
    
    return summary, site_status

# ============== Ruckus Import Functions ==============

def parse_ruckus_csv(file_content, offline_only=False):
    """
    Parse Ruckus CSV export (per-AP format).
    Aggregates APs by Zone and determines offline status based on LastSeen.
    
    Expected columns: MAC Address, AP Name, Zone, Domain, IP Address, IPv6 Address, 
                      IP Mode, Model, Mesh Role, LastSeen
    """
    lines = [line for line in file_content.splitlines() if line is not None]
    
    # Skip header rows (title rows before actual data)
    # Find the header row with column names
    header_idx = 0
    for i, line in enumerate(lines):
        # Match only the real header row, not metadata rows like "TimeZone:..."
        if 'MAC Address' in line and 'AP Name' in line:
            header_idx = i
            break
    
    # Parse from header onwards
    reader = csv.DictReader(lines[header_idx:])
    
    # Group APs by zone
    zones = {}
    now = datetime.utcnow()
    offline_threshold_hours = 8  # Consider AP offline if LastSeen > 8 hours ago
    
    for row in reader:
        zone_name = (row.get('Zone') or row.get('Zone Name') or '').strip()
        if not zone_name or zone_name == 'Zone':
            continue

        ap_name = (row.get('AP Name') or '').strip()
        status_value = (row.get('Status') or '').strip().lower()

        # Exclude known maintenance/decommissioned APs from scoring.
        ap_name_norm = ap_name.lower()
        is_maintenance = any(keyword in ap_name_norm for keyword in MAINTENANCE_AP_KEYWORDS)
        
        # Parse LastSeen to determine if AP is offline.
        # In offline-only mode, every row is treated as an offline AP by design.
        last_seen_str = row.get('LastSeen', '').strip()
        is_offline = True  # Conservative default
        source_basis = 'unknown'
        data_scope = 'offline_only' if offline_only else 'full_snapshot'

        if offline_only:
            is_offline = True
            source_basis = 'offline_only'
        else:
            # Prefer explicit status if the export provides it.
            if status_value:
                is_offline = status_value in {'offline', 'down', 'disconnected'}
                source_basis = 'status'
            elif last_seen_str:
                # Fall back to LastSeen age if there is no Status column.
                is_offline = True
                source_basis = 'last_seen'
                try:
                    # Format: "03/24/2026 10:30" or similar
                    for fmt in ['%m/%d/%Y %H:%M', '%Y-%m-%d %H:%M', '%d/%m/%Y %H:%M', '%m/%d/%Y %H:%M:%S']:
                        try:
                            last_seen = datetime.strptime(last_seen_str, fmt)
                            hours_ago = (now - last_seen).total_seconds() / 3600
                            is_offline = hours_ago > offline_threshold_hours
                            break
                        except ValueError:
                            continue
                except Exception:
                    is_offline = True
                    source_basis = 'unknown'
        
        # Initialize zone if not seen
        if zone_name not in zones:
            zones[zone_name] = {
                'total': 0,
                'offline': 0,
                'online': 0,
                'excluded_aps': 0,
                'excluded_examples': [],
                'offline_ap_examples': [],
                'basis_counts': {'status': 0, 'last_seen': 0, 'unknown': 0, 'offline_only': 0},
                'data_scope': data_scope
            }

        if is_maintenance:
            zones[zone_name]['excluded_aps'] += 1
            if len(zones[zone_name]['excluded_examples']) < 5 and ap_name:
                zones[zone_name]['excluded_examples'].append(ap_name)
            continue

        zones[zone_name]['total'] += 1
        zones[zone_name]['basis_counts'][source_basis] = zones[zone_name]['basis_counts'].get(source_basis, 0) + 1
        if is_offline:
            zones[zone_name]['offline'] += 1
            if ap_name and len(zones[zone_name]['offline_ap_examples']) < 25:
                zones[zone_name]['offline_ap_examples'].append(ap_name)
        else:
            zones[zone_name]['online'] += 1
    
    # Convert to results format
    results = []
    for zone_name, counts in zones.items():
        confidence = 'low'
        if counts['basis_counts'].get('unknown', 0) == 0:
            if counts['basis_counts'].get('status', 0) > 0 and counts['basis_counts'].get('last_seen', 0) == 0:
                confidence = 'high'
            elif counts['basis_counts'].get('status', 0) > 0 or counts['basis_counts'].get('last_seen', 0) > 0:
                confidence = 'medium'

        basis_used = []
        if counts['basis_counts'].get('status', 0):
            basis_used.append('status')
        if counts['basis_counts'].get('last_seen', 0):
            basis_used.append('last_seen')
        if counts['basis_counts'].get('unknown', 0):
            basis_used.append('unknown')

        results.append({
            'site_name': zone_name,
            'total_aps': counts['total'],
            'aps_online': counts['online'],
            'aps_offline': counts['offline'],
            'alerts': {
                'confidence': confidence,
                'source_basis': '+'.join(basis_used) if basis_used else 'unknown',
                'data_scope': counts.get('data_scope', 'full_snapshot'),
                'excluded_aps': counts['excluded_aps'],
                'excluded_examples': counts['excluded_examples'],
                'offline_ap_count': counts['offline'],
                'offline_ap_examples': counts['offline_ap_examples']
            }
        })
    
    return results

def parse_ruckus_json(file_content, offline_only=False):
    """Parse Ruckus JSON export."""
    data = json.loads(file_content)
    results = []
    
    # Handle different JSON structures
    zones = data if isinstance(data, list) else data.get('zones', data.get('data', data.get('list', [])))
    
    for zone in zones:
        site_name = zone.get('zoneName') or zone.get('name') or zone.get('zone_name') or zone.get('Zone Name') or ''
        total_aps = zone.get('totalAps') or zone.get('total_aps') or zone.get('apCount') or zone.get('Total APs') or 0
        aps_online = zone.get('onlineAps') or zone.get('aps_online') or zone.get('connectedAps') or zone.get('Online') or 0
        aps_offline = zone.get('offlineAps') or zone.get('aps_offline') or zone.get('disconnectedAps') or zone.get('Offline') or 0
        
        if not aps_offline and total_aps and aps_online:
            aps_offline = int(total_aps) - int(aps_online)
        
        if site_name:
            alerts_meta = zone.get('alerts', {})
            if not isinstance(alerts_meta, dict):
                alerts_meta = {}
            alerts_meta.setdefault('data_scope', 'offline_only' if offline_only else 'full_snapshot')
            alerts_meta.setdefault('offline_ap_count', int(aps_offline))
            results.append({
                'site_name': str(site_name).strip(),
                'total_aps': int(total_aps),
                'aps_online': int(aps_online),
                'aps_offline': int(aps_offline),
                'alerts': alerts_meta
            })
    
    return results

def match_zone_to_site(zone_name):
    """Match a Ruckus zone name to an assigned site."""
    zone_norm = normalize_site_name(zone_name)

    # Exact normalized alias match.
    if zone_norm in ALIAS_TO_SITE:
        return ALIAS_TO_SITE[zone_norm]

    # Fuzzy match for near-equivalent names (e.g., brackets, abbreviations, spacing).
    candidates = list(ALIAS_TO_SITE.keys())
    fuzzy = get_close_matches(zone_norm, candidates, n=1, cutoff=0.6)
    if fuzzy:
        return ALIAS_TO_SITE[fuzzy[0]]

    # Partial containment as a final fallback.
    for alias_norm, site_info in ALIAS_TO_SITE.items():
        if alias_norm and (alias_norm in zone_norm or zone_norm in alias_norm):
            return site_info

    return None

def import_ruckus_data(ruckus_data, offline_only=False):
    """Import parsed Ruckus data, filtering to assigned sites only."""
    
    # Aggregate data by site (multiple zones may belong to same site)
    site_aggregates = {}
    records_processed = len(ruckus_data)
    unmatched_zones = set()
    
    for data in ruckus_data:
        zone_name = data['site_name']
        site_info = match_zone_to_site(zone_name)
        
        if site_info:
            site_name = site_info['name']
            region = site_info['region']
            
            if site_name not in site_aggregates:
                site_aggregates[site_name] = {
                    'region': region,
                    'total_aps': 0,
                    'aps_online': 0,
                    'aps_offline': 0,
                    'zones_matched': []
                }
            
            site_aggregates[site_name]['total_aps'] += data['total_aps']
            site_aggregates[site_name]['aps_online'] += data['aps_online']
            site_aggregates[site_name]['aps_offline'] += data['aps_offline']
            site_aggregates[site_name]['zones_matched'].append(zone_name)
            site_aggregates[site_name].setdefault('excluded_aps', 0)
            site_aggregates[site_name].setdefault('excluded_examples', [])
            site_aggregates[site_name].setdefault('offline_ap_examples', [])
            site_aggregates[site_name].setdefault('basis_used', set())
            site_aggregates[site_name].setdefault('data_scopes', set())

            alerts_meta = data.get('alerts') if isinstance(data.get('alerts'), dict) else {}
            site_aggregates[site_name]['excluded_aps'] += int(alerts_meta.get('excluded_aps', 0) or 0)
            for ex in alerts_meta.get('excluded_examples', []) or []:
                if ex not in site_aggregates[site_name]['excluded_examples'] and len(site_aggregates[site_name]['excluded_examples']) < 8:
                    site_aggregates[site_name]['excluded_examples'].append(ex)

            for ap in alerts_meta.get('offline_ap_examples', []) or []:
                if ap not in site_aggregates[site_name]['offline_ap_examples'] and len(site_aggregates[site_name]['offline_ap_examples']) < 50:
                    site_aggregates[site_name]['offline_ap_examples'].append(ap)

            basis = alerts_meta.get('source_basis')
            if basis:
                for part in str(basis).split('+'):
                    if part:
                        site_aggregates[site_name]['basis_used'].add(part)

            data_scope = alerts_meta.get('data_scope')
            if data_scope:
                site_aggregates[site_name]['data_scopes'].add(str(data_scope))
        else:
            unmatched_zones.add(zone_name)
    
    records_matched = 0
    details = []
    
    # Now create/update sites and monitoring logs
    for site_name, agg_data in site_aggregates.items():
        records_matched += 1
        
        # Get or create site in database
        site = Site.query.filter_by(name=site_name).first()
        if not site:
            site = Site(
                name=site_name,
                region=agg_data['region'],
                ruckus_zone_name=', '.join(agg_data['zones_matched'][:3]),  # Store first few matching zones
                total_aps=agg_data['total_aps'],
                active=True
            )
            db.session.add(site)
            db.session.flush()  # Get the ID
        else:
            if not offline_only:
                site.total_aps = agg_data['total_aps']

        # Build effective totals.
        effective_total_aps = agg_data['total_aps']
        effective_online_aps = agg_data['aps_online']
        status_note = ''

        if offline_only:
            baseline_total = 0
            if site.total_aps and site.total_aps > 0:
                baseline_total = site.total_aps
            else:
                previous = MonitoringLog.query.filter_by(site_id=site.id)\
                    .order_by(MonitoringLog.timestamp.desc()).first()
                if previous and previous.total_aps and previous.total_aps > 0:
                    baseline_total = previous.total_aps

            if baseline_total >= agg_data['aps_offline'] and baseline_total > 0:
                effective_total_aps = baseline_total
                effective_online_aps = max(0, baseline_total - agg_data['aps_offline'])
                status_note = 'offline_only_with_baseline'
            else:
                effective_total_aps = agg_data['aps_offline']
                effective_online_aps = 0
                status_note = 'offline_only_no_baseline'
        
        # Calculate status
        if offline_only and status_note == 'offline_only_no_baseline':
            status = 'unknown'
        else:
            status = calculate_status(effective_total_aps, agg_data['aps_offline'])

        confidence = 'low'
        if 'unknown' not in agg_data['basis_used']:
            if agg_data['basis_used'] == {'status'}:
                confidence = 'high'
            elif agg_data['basis_used']:
                confidence = 'medium'

        if offline_only and confidence == 'low':
            confidence = 'medium'

        previous_log = MonitoringLog.query.filter_by(site_id=site.id)\
            .order_by(MonitoringLog.timestamp.desc()).first()
        previous_meta = parse_alert_metadata(previous_log.alerts) if previous_log else {}
        previous_offline = set(previous_meta.get('offline_ap_examples', []) or [])
        current_offline = agg_data['offline_ap_examples']
        new_offline_examples = [ap for ap in current_offline if ap not in previous_offline][:50]
        restored_count = len([ap for ap in previous_offline if ap not in set(current_offline)]) if previous_offline else 0
        
        # Create monitoring log
        log = MonitoringLog(
            site_id=site.id,
            total_aps=effective_total_aps,
            aps_online=effective_online_aps,
            aps_offline=agg_data['aps_offline'],
            status=status,
            alerts=json.dumps({
                'confidence': confidence,
                'source_basis': '+'.join(sorted(agg_data['basis_used'])) if agg_data['basis_used'] else 'unknown',
                'data_scope': '+'.join(sorted(agg_data['data_scopes'])) if agg_data['data_scopes'] else ('offline_only' if offline_only else 'full_snapshot'),
                'excluded_aps': agg_data['excluded_aps'],
                'excluded_examples': agg_data['excluded_examples'],
                'offline_ap_count': agg_data['aps_offline'],
                'offline_ap_examples': agg_data['offline_ap_examples'],
                'new_offline_count': len(new_offline_examples),
                'new_offline_examples': new_offline_examples,
                'restored_count': restored_count,
                'status_note': status_note
            }),
            auto_imported=True
        )
        db.session.add(log)
        
        if agg_data['aps_offline'] > 0:
            details.append(f"{site_name}: {agg_data['aps_offline']}/{agg_data['total_aps']} offline ({status})")
    
    db.session.commit()
    
    # Log the import
    status = 'success'
    if records_matched == 0:
        status = 'failed'
    elif unmatched_zones:
        status = 'partial'

    unmatched_list = sorted(unmatched_zones)
    unmatched_preview = ', '.join(unmatched_list[:10])
    unmatched_note = f"Unmatched zones ({len(unmatched_list)}): {unmatched_preview}" if unmatched_list else ''

    import_log = ImportLog(
        source='ruckus',
        records_processed=records_processed,
        records_matched=records_matched,
        status=status,
        details='\n'.join([msg for msg in [
            ('; '.join(details) if details else f'All {records_matched} matched sites healthy'),
            unmatched_note
        ] if msg])
    )
    db.session.add(import_log)
    db.session.commit()
    
    return {
        'processed': records_processed,
        'matched': records_matched,
        'ignored': len(unmatched_list),
        'unmatched_sites': unmatched_list,
        'issues': details,
        'sites': list(site_aggregates.keys())
    }

# ============== Freshdesk Integration ==============

def _freshdesk_base_url():
    """Return a clean Freshdesk base URL, handling full URLs and trailing slashes."""
    domain = (get_config('freshdesk_domain') or '').strip().rstrip('/').rstrip('\\').strip()
    if not domain:
        return None
    domain = re.sub(r'^https?://', '', domain)
    domain = domain.strip('/')
    if '.' not in domain:
        domain = f'{domain}.freshdesk.com'
    return f'https://{domain}'


def resolve_self_freshdesk_agent_id():
    """Auto-discover and cache the numeric Freshdesk agent ID for the configured API key."""
    cached = get_config('freshdesk_self_agent_id')
    if cached:
        return cached, None

    base = _freshdesk_base_url()
    api_key = (get_config('freshdesk_api_key') or '').strip()
    if not base or not api_key:
        return None, 'Freshdesk not configured'

    try:
        resp = requests.get(f'{base}/api/v2/agents/me', auth=(api_key, 'X'), timeout=15)
        if resp.status_code == 200:
            numeric_id = str(resp.json().get('id', ''))
            if numeric_id:
                set_config('freshdesk_self_agent_id', numeric_id)
                return numeric_id, None
        return None, f'Could not resolve agent ID (HTTP {resp.status_code}: {resp.text[:200]})'
    except Exception as e:
        return None, str(e)


def fetch_freshdesk_tickets(agent_freshdesk_id, start_date, end_date):
    """Fetch tickets from Freshdesk API."""
    api_key = (get_config('freshdesk_api_key') or '').strip()
    base = _freshdesk_base_url()

    if not api_key or not base:
        return None, 'Freshdesk not configured'

    search_url = f'{base}/api/v2/search/tickets'

    try:
        # --- Query 1: tickets OPENED (created) in the date range ---
        created_query = f"\"agent_id:{agent_freshdesk_id} AND created_at:>'{start_date}' AND created_at:<'{end_date}'\""
        r1 = requests.get(search_url, auth=(api_key, 'X'), params={'query': created_query}, timeout=30)
        if r1.status_code != 200:
            return None, f'API error {r1.status_code}: {r1.text[:200]}'
        created_tickets = r1.json().get('results', [])
        handled = len(created_tickets)

        # --- Query 2a: tickets RESOLVED in the date range (regardless of when created) ---
        resolved_query = f"\"agent_id:{agent_freshdesk_id} AND resolved_at:>'{start_date}' AND resolved_at:<'{end_date}'\""
        r_resolved = requests.get(search_url, auth=(api_key, 'X'), params={'query': resolved_query}, timeout=30)
        resolved_count = r_resolved.json().get('total', 0) if r_resolved.status_code == 200 else 0
        closed = resolved_count

        # --- Query 2: currently OPEN tickets (status=2) for this agent ---
        r_open = requests.get(search_url, auth=(api_key, 'X'),
            params={'query': f'"agent_id:{agent_freshdesk_id} AND status:2"'}, timeout=30)
        open_count = r_open.json().get('total', 0) if r_open.status_code == 200 else 0

        # --- Query 3: currently PENDING tickets (status=3) for this agent ---
        r_pend = requests.get(search_url, auth=(api_key, 'X'),
            params={'query': f'"agent_id:{agent_freshdesk_id} AND status:3"'}, timeout=30)
        pending_count = r_pend.json().get('total', 0) if r_pend.status_code == 200 else 0

        total_pending = open_count + pending_count

        # Escalated = urgent/high priority among currently open+pending tickets
        open_results = r_open.json().get('results', []) if r_open.status_code == 200 else []
        pend_results = r_pend.json().get('results', []) if r_pend.status_code == 200 else []
        escalated = sum(1 for t in open_results + pend_results if t.get('priority') in [3, 4])

        return {
            'handled': handled,
            'closed': closed,
            'escalated': escalated,
            'pending': total_pending
        }, None

    except Exception as e:
        return None, str(e)


def sync_freshdesk_for_agent(agent_id, date=None):
    """Sync Freshdesk data for a specific agent."""
    agent = Agent.query.get(agent_id)
    if not agent:
        return False, 'Agent not found'

    # Use stored agent ID, or auto-discover from the API key owner.
    fd_agent_id = agent.freshdesk_agent_id
    if not fd_agent_id:
        fd_agent_id, err = resolve_self_freshdesk_agent_id()
        if not fd_agent_id:
            return False, err or 'Could not resolve Freshdesk agent ID'
        # Persist so next call is instant.
        agent.freshdesk_agent_id = fd_agent_id
        db.session.commit()

    if date is None:
        date = datetime.now().date()

    data, error = fetch_freshdesk_tickets(
        fd_agent_id,
        date.isoformat(),
        (date + timedelta(days=1)).isoformat()
    )

    if data:
        stats = TicketStats.query.filter_by(agent_id=agent.id, date=date).first()

        if stats:
            stats.tickets_handled = data['handled']
            stats.tickets_closed = data['closed']
            stats.tickets_escalated = data['escalated']
            stats.tickets_pending = data['pending']
            stats.auto_imported = True
            stats.source = 'freshdesk'
        else:
            stats = TicketStats(
                agent_id=agent.id,
                date=date,
                tickets_handled=data['handled'],
                tickets_closed=data['closed'],
                tickets_escalated=data['escalated'],
                tickets_pending=data['pending'],
                auto_imported=True,
                source='freshdesk'
            )
            db.session.add(stats)

        db.session.commit()
        return True, data

    return False, error

# ============== Report Generation ==============

def generate_weekly_report_auto(agent_id, week_ending):
    """Auto-generate weekly report from collected data."""
    week_start = get_week_start(week_ending)
    
    report = WeeklyReport.query.filter_by(agent_id=agent_id, week_ending=week_ending).first()
    
    if not report:
        report = WeeklyReport(agent_id=agent_id, week_ending=week_ending)
        db.session.add(report)
    
    # Count site visits
    visits = SiteVisit.query.filter(
        SiteVisit.agent_id == agent_id,
        SiteVisit.date >= week_start,
        SiteVisit.date <= week_ending
    ).count()
    report.sites_visited = visits
    
    # Count monitored sites (unique sites with logs this week)
    week_start_dt = datetime.combine(week_start, datetime.min.time())
    week_end_dt = datetime.combine(week_ending, datetime.max.time())
    
    monitoring_metrics = calculate_weekly_monitoring_metrics(week_start_dt, week_end_dt)
    report.sites_monitored = monitoring_metrics['sites_monitored']
    report.aps_offline_total = monitoring_metrics['aps_offline_total']
    report.critical_incidents = monitoring_metrics['critical_incidents']
    
    # Sum ticket stats
    tickets = TicketStats.query.filter(
        TicketStats.agent_id == agent_id,
        TicketStats.date >= week_start,
        TicketStats.date <= week_ending
    ).all()
    
    report.tickets_handled = sum(t.tickets_handled for t in tickets)
    report.tickets_closed = sum(t.tickets_closed for t in tickets)
    report.auto_generated = True
    
    db.session.commit()
    return report

# ============== Routes ==============

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '')
        user = User.query.filter_by(email=email).first()
        if user and user.active and user.check_password(password):
            from flask import session as flask_session
            flask_session.permanent = True
            user.last_login = datetime.utcnow()
            db.session.commit()
            login_user(user)
            next_page = request.args.get('next')
            if next_page and next_page.startswith('/'):
                return redirect(next_page)
            return redirect(url_for('index'))
        flash('Invalid email or password.', 'danger')
    return render_template('login.html')


@app.route('/logout')
def logout():
    logout_user()
    return redirect(url_for('login'))


@app.route('/')
def index():
    """Dashboard with health overview."""
    agent_id = get_active_agent_id()

    # No-agent warning for viewers
    if current_user.is_authenticated and not current_user.is_admin and not agent_id:
        flash('No agent is assigned to your account. Please contact an administrator.', 'warning')

    health_summary, site_status = get_site_health_summary()
    agents = Agent.query.filter_by(active=True).all()
    week_ending = get_week_ending()
    
    # Recent imports
    recent_imports = ImportLog.query.order_by(ImportLog.timestamp.desc()).limit(5).all()
    
    # Today's ticket stats — scoped
    today = datetime.now().date()
    ts_query = TicketStats.query.filter_by(date=today)
    if agent_id:
        ts_query = ts_query.filter_by(agent_id=agent_id)
    today_stats = ts_query.all()
    total_tickets_today = sum(s.tickets_handled for s in today_stats)
    
    # Sites needing attention (critical/warning)
    attention_sites = [s for s in site_status if s['status'] in ['warning', 'critical']]

    # Data quality and maintenance visibility for one-stop reporting
    confidence_counts = {'high': 0, 'medium': 0, 'low': 0}
    for s in site_status:
        level = s.get('confidence', 'low')
        confidence_counts[level] = confidence_counts.get(level, 0) + 1

    total_excluded_aps = sum(s.get('excluded_aps', 0) for s in site_status)
    maintenance_sites = [
        {
            'site': s['site'].name,
            'excluded_aps': s.get('excluded_aps', 0),
            'examples': s.get('excluded_examples', [])
        }
        for s in site_status if s.get('excluded_aps', 0) > 0
    ]
    maintenance_sites.sort(key=lambda x: x['excluded_aps'], reverse=True)

    last_ruckus_import = ImportLog.query.filter_by(source='ruckus')\
        .order_by(ImportLog.timestamp.desc()).first()

    unmatched_count = 0
    if last_ruckus_import and last_ruckus_import.details:
        match = re.search(r'Unmatched zones \((\d+)\)', last_ruckus_import.details)
        if match:
            unmatched_count = int(match.group(1))

    data_quality = {
        'last_import': last_ruckus_import.timestamp if last_ruckus_import else None,
        'import_status': last_ruckus_import.status if last_ruckus_import else 'no_data',
        'matched': last_ruckus_import.records_matched if last_ruckus_import else 0,
        'processed': last_ruckus_import.records_processed if last_ruckus_import else 0,
        'unmatched': unmatched_count,
        'confidence_counts': confidence_counts,
        'total_excluded_aps': total_excluded_aps
    }
    
    approved_today = LeaveRequest.query.filter(
        LeaveRequest.status == 'approved',
        LeaveRequest.start_date <= today,
        LeaveRequest.end_date >= today
    ).count()

    _we = get_week_ending()
    _ws = _we - timedelta(days=4)
    return render_template('index.html',
        health_summary=health_summary,
        site_status=site_status,
        attention_sites=attention_sites,
        agents=agents,
        week_ending=week_ending,
        recent_imports=recent_imports,
        total_tickets_today=total_tickets_today,
        data_quality=data_quality,
        maintenance_sites=maintenance_sites,
        leave_pending=LeaveRequest.query.filter_by(status='pending').count(),
        leave_approved=LeaveRequest.query.filter_by(status='approved').filter(
            LeaveRequest.end_date >= today).count(),
        escalated_this_week=db.session.query(db.func.sum(TicketStats.tickets_escalated)).filter(
            TicketStats.date >= _ws).scalar() or 0,
        tickets_closed_this_week=db.session.query(db.func.sum(TicketStats.tickets_closed)).filter(
            TicketStats.date >= _ws).scalar() or 0,
        total_aps_offline=sum(s.get('aps_offline', 0) for s in site_status),
        total_sites=len([s for s in site_status if s['status'] != 'unknown']),
        total_engineers=len(agents),
        engineers_available=max(len(agents) - approved_today, 0),
        ticket_trend=[
            {
                'week': (_we - timedelta(weeks=i)).strftime('%d %b'),
                'closed': db.session.query(db.func.sum(TicketStats.tickets_closed)).filter(
                    TicketStats.date >= (_we - timedelta(weeks=i, days=4)),
                    TicketStats.date <= (_we - timedelta(weeks=i))
                ).scalar() or 0
            }
            for i in range(3, -1, -1)
        ]
    )

# ----- Site Health -----

@app.route('/health')
def site_health():
    """Detailed site health view."""
    health_summary, site_status = get_site_health_summary()
    
    # Sort by status (critical first)
    status_order = {'critical': 0, 'warning': 1, 'unknown': 2, 'healthy': 3}
    site_status.sort(key=lambda x: status_order.get(x['status'], 4))
    
    return render_template('health.html',
        health_summary=health_summary,
        site_status=site_status
    )

# ----- Site Management -----

@app.route('/sites')
def sites():
    show_inactive = request.args.get('show_inactive') == '1'
    all_sites = Site.query.order_by(Site.region, Site.name).all()
    active_sites = [s for s in all_sites if s.active]
    inactive_sites = [s for s in all_sites if not s.active]
    display_sites = all_sites if show_inactive else active_sites
    within_reach = [s for s in display_sites if s.region == 'within_reach']
    far_reach = [s for s in display_sites if s.region == 'far_reach']
    return render_template('sites.html',
        within_reach=within_reach,
        far_reach=far_reach,
        all_sites=all_sites,
        show_inactive=show_inactive,
        inactive_count=len(inactive_sites)
    )

@app.route('/sites/add', methods=['GET', 'POST'])
def add_site():
    if request.method == 'POST':
        site = Site(
            name=request.form['name'],
            region=request.form['region'],
            ruckus_zone_name=request.form.get('ruckus_zone_name') or request.form['name'],
            total_aps=int(request.form.get('total_aps', 0)),
            active=True
        )
        db.session.add(site)
        db.session.commit()
        flash(f'Site "{site.name}" added.', 'success')
        return redirect(url_for('sites'))
    return render_template('add_site.html')

@app.route('/sites/import', methods=['GET', 'POST'])
def import_sites():
    """Bulk import sites from JSON."""
    if request.method == 'POST':
        try:
            data = json.loads(request.form['sites_json'])
            count = 0
            
            for region, site_list in data.items():
                for site_name in site_list:
                    existing = Site.query.filter_by(name=site_name).first()
                    if not existing:
                        site = Site(
                            name=site_name,
                            region=region,
                            ruckus_zone_name=site_name
                        )
                        db.session.add(site)
                        count += 1
            
            db.session.commit()
            flash(f'Imported {count} new sites.', 'success')
            return redirect(url_for('sites'))
        except Exception as e:
            flash(f'Error: {str(e)}', 'danger')
    
    return render_template('import_sites.html')

@app.route('/sites/<int:id>/toggle')
def toggle_site(id):
    site = Site.query.get_or_404(id)
    site.active = not site.active
    db.session.commit()
    flash(f'Site "{site.name}" {"activated" if site.active else "deactivated"}.', 'info')
    return redirect(url_for('sites'))

@app.route('/sites/<int:id>/delete', methods=['POST'])
def delete_site(id):
    site = Site.query.get_or_404(id)
    name = site.name
    db.session.delete(site)
    db.session.commit()
    flash(f'Site "{name}" deleted.', 'success')
    return redirect(url_for('sites'))

# ----- Ruckus Import -----

@app.route('/import/ruckus', methods=['GET', 'POST'])
def import_ruckus():
    """Import Ruckus data from CSV/JSON."""
    if request.method == 'POST':
        offline_only = request.form.get('offline_only') == '1'

        if 'ruckus_file' in request.files and request.files['ruckus_file'].filename:
            file = request.files['ruckus_file']
            content = file.read().decode('utf-8')
            
            if file.filename.endswith('.json'):
                data = parse_ruckus_json(content, offline_only=offline_only)
            else:
                data = parse_ruckus_csv(content, offline_only=offline_only)
            
            result = import_ruckus_data(data, offline_only=offline_only)
            flash(
                f'Imported: {result["matched"]} sites matched. '
                f'Ignored: {result.get("ignored", 0)} unmatched zones.',
                'success' if result['matched'] > 0 else 'warning'
            )
            if offline_only:
                flash('Import mode: Offline APs only. Baseline totals are reused when available.', 'info')
            if result.get('unmatched_sites'):
                preview = ', '.join(result['unmatched_sites'][:10])
                flash(f'Unmatched Zones: {preview}', 'warning')
            if result['issues']:
                flash(f'Issues found: {len(result["issues"])}', 'warning')
            
            return redirect(url_for('site_health'))
        
        elif request.form.get('paste_data'):
            content = request.form['paste_data'].strip()
            offline_only = request.form.get('offline_only') == '1'
            try:
                data = parse_ruckus_json(content, offline_only=offline_only)
            except:
                data = parse_ruckus_csv(content, offline_only=offline_only)
            
            result = import_ruckus_data(data, offline_only=offline_only)
            flash(
                f'Imported: {result["matched"]} sites matched. '
                f'Ignored: {result.get("ignored", 0)} unmatched zones.',
                'success' if result['matched'] > 0 else 'warning'
            )
            if offline_only:
                flash('Import mode: Offline APs only. Baseline totals are reused when available.', 'info')
            if result.get('unmatched_sites'):
                preview = ', '.join(result['unmatched_sites'][:10])
                flash(f'Unmatched Zones: {preview}', 'warning')
            return redirect(url_for('site_health'))
    
    recent_imports = ImportLog.query.filter_by(source='ruckus')\
        .order_by(ImportLog.timestamp.desc()).limit(10).all()
    
    return render_template('import_ruckus.html', recent_imports=recent_imports)

@app.route('/import/ruckus/clear', methods=['POST'])
def clear_ruckus_data():
    """Delete all monitoring logs and Ruckus import history to start fresh."""
    ml_count = MonitoringLog.query.count()
    MonitoringLog.query.delete()
    ImportLog.query.filter_by(source='ruckus').delete()
    db.session.commit()
    flash(f'Cleared {ml_count} monitoring log(s). Ready for a fresh import.', 'success')
    return redirect(url_for('import_ruckus'))

# ----- Agent Management -----

@app.route('/agents')
@admin_required
def agents():
    all_agents = Agent.query.all()
    return render_template('agents.html', agents=all_agents)

@app.route('/agents/add', methods=['GET', 'POST'])
@admin_required
def add_agent():
    if request.method == 'POST':
        agent = Agent(
            name=request.form['name'],
            email=request.form.get('email', ''),
            freshdesk_agent_id=request.form.get('freshdesk_agent_id'),
            active=True
        )
        db.session.add(agent)
        db.session.commit()
        flash(f'Agent {agent.name} added.', 'success')
        return redirect(url_for('agents'))
    return render_template('add_agent.html')

@app.route('/agents/<int:id>/toggle')
@admin_required
def toggle_agent(id):
    agent = Agent.query.get_or_404(id)
    agent.active = not agent.active
    db.session.commit()
    return redirect(url_for('agents'))

@app.route('/agents/<int:id>/delete', methods=['POST'])
@admin_required
def delete_agent(id):
    agent = Agent.query.get_or_404(id)
    name = agent.name
    # Cascade-delete all associated data before removing the agent
    SiteVisit.query.filter_by(agent_id=id).delete()
    TicketStats.query.filter_by(agent_id=id).delete()
    WeeklyReport.query.filter_by(agent_id=id).delete()
    ServiceCall.query.filter_by(agent_id=id).delete()
    db.session.delete(agent)
    db.session.commit()
    flash(f'Agent "{name}" and all their data deleted.', 'success')
    return redirect(url_for('agents'))

# ----- Site Visits (Manual) -----

@app.route('/site-visits')
def site_visits():
    agent_id = get_active_agent_id()
    query = SiteVisit.query
    if agent_id:
        query = query.filter_by(agent_id=agent_id)
    visits = query.order_by(SiteVisit.date.desc()).limit(50).all()
    return render_template('site_visits.html', visits=visits)

@app.route('/site-visits/add', methods=['GET', 'POST'])
def add_site_visit():
    agents = Agent.query.filter_by(active=True).all()
    all_sites = Site.query.filter_by(active=True).order_by(Site.name).all()

    if request.method == 'POST':
        site_id = request.form.get('site_id')
        site_name = request.form.get('site_name', '').strip()
        # 'other' means user typed a custom name; numeric means a known site
        if site_id and site_id != 'other':
            try:
                site_obj = Site.query.get(int(site_id))
                site_name = site_obj.name if site_obj else site_name
            except (ValueError, TypeError):
                site_id = None
        else:
            site_id = None

        visit = SiteVisit(
            agent_id=request.form['agent_id'],
            site_id=int(site_id) if site_id else None,
            site_name=site_name,
            location=request.form.get('location', '').strip() or None,
            date=datetime.strptime(request.form['date'], '%Y-%m-%d').date(),
            discussion_topics=request.form.get('discussion_topics', ''),
            info_obtained=request.form.get('info_obtained', ''),
            follow_up_actions=request.form.get('follow_up_actions', '')
        )
        db.session.add(visit)
        db.session.commit()
        # Keep the weekly report summary in sync
        week_ending = get_week_ending(datetime.combine(visit.date, datetime.min.time()))
        generate_weekly_report_auto(visit.agent_id, week_ending)
        flash('Site visit logged.', 'success')
        return redirect(url_for('site_visits'))

    return render_template('add_site_visit.html',
        agents=agents, sites=all_sites, today=datetime.now().strftime('%Y-%m-%d'))

@app.route('/service-calls')
def service_calls():
    agent_id = get_active_agent_id()
    query = ServiceCall.query
    if agent_id:
        query = query.filter_by(agent_id=agent_id)
    calls = query.order_by(ServiceCall.date.desc()).limit(100).all()
    return render_template('service_calls.html', calls=calls)

@app.route('/service-calls/add', methods=['GET', 'POST'])
def add_service_call():
    agents = Agent.query.filter_by(active=True).all()
    if request.method == 'POST':
        call = ServiceCall(
            agent_id=request.form['agent_id'],
            date=datetime.strptime(request.form['date'], '%Y-%m-%d').date(),
            ticket_number=request.form.get('ticket_number', '').strip() or None,
            summary=request.form.get('summary', ''),
            status=request.form.get('status', 'open'),
            resolution=request.form.get('resolution', ''),
            time_spent=int(request.form['time_spent']) if request.form.get('time_spent') else None
        )
        db.session.add(call)
        db.session.commit()
        flash('Service call logged.', 'success')
        return redirect(url_for('service_calls'))
    return render_template('add_service_call.html', agents=agents, today=datetime.now().strftime('%Y-%m-%d'))

@app.route('/service-calls/bulk', methods=['GET', 'POST'])
def bulk_service_calls():
    agents = Agent.query.filter_by(active=True).all()
    if request.method == 'POST':
        date = datetime.strptime(request.form['date'], '%Y-%m-%d').date()
        agent_id = request.form['agent_id']
        existing = TicketStats.query.filter_by(agent_id=agent_id, date=date).first()
        if existing:
            existing.tickets_handled = int(request.form.get('tickets_handled', 0))
            existing.tickets_closed = int(request.form.get('tickets_closed', 0))
            existing.source = 'manual'
        else:
            stats = TicketStats(
                agent_id=agent_id,
                date=date,
                tickets_handled=int(request.form.get('tickets_handled', 0)),
                tickets_closed=int(request.form.get('tickets_closed', 0)),
                auto_imported=False,
                source='manual'
            )
            db.session.add(stats)
        db.session.commit()
        flash('Ticket counts saved.', 'success')
        return redirect(url_for('service_calls'))
    return render_template('bulk_service_calls.html', agents=agents, today=datetime.now().strftime('%Y-%m-%d'))

# ----- Monitoring -----

@app.route('/monitoring')
def monitoring():
    """View auto-imported monitoring data."""
    selected_site_id = request.args.get('site_id', type=int)
    selected_status = (request.args.get('status') or '').strip().lower()
    ap_query = (request.args.get('ap_query') or '').strip().lower()
    only_offline = request.args.get('only_offline', '1') != '0'

    # Get all sites for filter
    all_sites = Site.query.filter_by(active=True).order_by(Site.name).all()
    
    # Get latest log per site
    subq = db.session.query(
        MonitoringLog.site_id,
        db.func.max(MonitoringLog.timestamp).label('max_time')
    ).group_by(MonitoringLog.site_id).subquery()
    
    latest_logs = db.session.query(MonitoringLog).join(
        subq,
        db.and_(
            MonitoringLog.site_id == subq.c.site_id,
            MonitoringLog.timestamp == subq.c.max_time
        )
    ).all()

    log_ids = [log.id for log in latest_logs]
    escalation_rows = []
    if log_ids:
        escalation_rows = MonitoringEscalation.query.filter(
            MonitoringEscalation.monitoring_log_id.in_(log_ids)
        ).order_by(MonitoringEscalation.escalated_at.desc()).all()
    escalation_map = {}
    for row in escalation_rows:
        if row.monitoring_log_id not in escalation_map:
            escalation_map[row.monitoring_log_id] = row
    
    # Attach parsed metadata for template rendering and filter records.
    filtered_logs = []
    for log in latest_logs:
        meta = parse_alert_metadata(log.alerts)
        log.import_confidence = meta['confidence']
        log.import_basis = meta['source_basis']
        log.excluded_aps = meta['excluded_aps']
        log.excluded_examples = meta['excluded_examples']
        log.offline_ap_count = meta['offline_ap_count']
        log.data_scope = meta['data_scope']
        log.new_offline_count = meta['new_offline_count']
        log.new_offline_examples = meta['new_offline_examples']
        log.restored_count = meta['restored_count']
        log.status_note = meta['status_note']
        # Keep list order from import while removing duplicates.
        seen_names = set()
        log.offline_ap_examples = []
        for ap_name in (meta.get('offline_ap_examples') or []):
            ap_name_clean = str(ap_name).strip()
            if ap_name_clean and ap_name_clean not in seen_names:
                seen_names.add(ap_name_clean)
                log.offline_ap_examples.append(ap_name_clean)
        log.offline_ap_text = '\n'.join(log.offline_ap_examples)
        log.latest_escalation = escalation_map.get(log.id)

        if selected_site_id and log.site_id != selected_site_id:
            continue
        if selected_status and (log.status or '').lower() != selected_status:
            continue
        if only_offline and log.aps_offline <= 0:
            continue
        if ap_query:
            if not any(ap_query in ap.lower() for ap in log.offline_ap_examples):
                continue

        filtered_logs.append(log)

    # Sort by status (critical first), then by site name.
    status_order = {'critical': 0, 'warning': 1, 'unknown': 2, 'healthy': 3}
    filtered_logs.sort(key=lambda x: (status_order.get(x.status, 4), (x.site.name if x.site else '')))

    return render_template(
        'monitoring.html',
        logs=filtered_logs,
        sites=all_sites,
        escalation_targets=ESCALATION_TARGETS
    )


@app.route('/monitoring/<int:id>/escalate', methods=['POST'])
def escalate_monitoring_issue(id):
    log = MonitoringLog.query.get_or_404(id)
    escalated_to = (request.form.get('escalated_to') or '').strip()
    escalation_reason = (request.form.get('escalation_reason') or '').strip()

    if escalated_to not in ESCALATION_TARGETS:
        flash('Please select a valid escalation target.', 'warning')
        return redirect(url_for('monitoring'))

    message = (
        f'[Escalation] Site: {log.site.name if log.site else "Unknown"}\n'
        f'Status: {log.status}\n'
        f'Offline APs: {log.aps_offline}/{log.total_aps}\n'
        f'Escalated to: {escalated_to}\n'
        f'Reason: {escalation_reason or "No reason provided"}'
    )
    tg_sent, tg_error = send_telegram_alert(message)

    escalation = MonitoringEscalation(
        monitoring_log_id=log.id,
        escalated=True,
        escalated_to=escalated_to,
        escalation_reason=escalation_reason,
        telegram_sent=tg_sent
    )
    db.session.add(escalation)
    db.session.commit()

    if tg_sent:
        flash(f'Issue escalated to {escalated_to}. Telegram alert sent.', 'success')
    else:
        flash(f'Issue escalated to {escalated_to}. Telegram not sent ({tg_error}).', 'warning')
    return redirect(url_for('monitoring'))


@app.route('/monitoring/escalations/<int:id>/resolve', methods=['POST'])
def resolve_monitoring_escalation(id):
    escalation = MonitoringEscalation.query.get_or_404(id)
    escalation.resolution_notes = (request.form.get('resolution_notes') or '').strip()
    escalation.resolved_at = datetime.utcnow()
    db.session.commit()
    flash('Escalation marked as resolved.', 'success')
    return redirect(url_for('monitoring'))

@app.route('/monitoring/<int:id>/note', methods=['GET', 'POST'])
def add_monitoring_note(id):
    log = MonitoringLog.query.get_or_404(id)
    
    if request.method == 'POST':
        log.notes = request.form.get('notes', '')
        db.session.commit()
        flash('Note added.', 'success')
        return redirect(url_for('monitoring'))
    
    return render_template('add_monitoring_note.html', log=log)

# ----- Tickets -----

@app.route('/tickets')
def tickets():
    agent_id = get_active_agent_id()
    stats_query = TicketStats.query
    if agent_id:
        stats_query = stats_query.filter_by(agent_id=agent_id)
    stats = stats_query.order_by(TicketStats.date.desc()).limit(50).all()
    
    # Weekly totals — scoped
    week_ending = get_week_ending()
    week_start = get_week_start(week_ending)
    wk_query = TicketStats.query.filter(
        TicketStats.date >= week_start,
        TicketStats.date <= week_ending
    )
    if agent_id:
        wk_query = wk_query.filter_by(agent_id=agent_id)
    weekly_stats = wk_query.all()
    # pending is a live snapshot - use the most recent value this week
    latest_with_pending = next(
        (s for s in sorted(weekly_stats, key=lambda s: s.date, reverse=True)
         if s.tickets_pending is not None), None
    )
    weekly_total = {
        'opened': sum(s.tickets_handled for s in weekly_stats),
        'closed': sum(s.tickets_closed for s in weekly_stats),
        'pending': latest_with_pending.tickets_pending if latest_with_pending else 0,
        'escalated': sum(s.tickets_escalated for s in weekly_stats)
    }
    
    freshdesk_configured = bool(get_config('freshdesk_api_key'))
    
    return render_template('tickets.html', 
        stats=stats, 
        weekly_total=weekly_total,
        freshdesk_configured=freshdesk_configured)

@app.route('/tickets/add', methods=['GET', 'POST'])
def add_tickets():
    agents = Agent.query.filter_by(active=True).all()
    
    if request.method == 'POST':
        date = datetime.strptime(request.form['date'], '%Y-%m-%d').date()
        agent_id = request.form['agent_id']
        
        existing = TicketStats.query.filter_by(agent_id=agent_id, date=date).first()
        
        if existing:
            existing.tickets_handled = int(request.form.get('tickets_handled', 0))
            existing.tickets_closed = int(request.form.get('tickets_closed', 0))
            existing.tickets_escalated = int(request.form.get('tickets_escalated', 0))
            existing.source = 'manual'
        else:
            stats = TicketStats(
                agent_id=agent_id,
                date=date,
                tickets_handled=int(request.form.get('tickets_handled', 0)),
                tickets_closed=int(request.form.get('tickets_closed', 0)),
                tickets_escalated=int(request.form.get('tickets_escalated', 0)),
                auto_imported=False,
                source='manual'
            )
            db.session.add(stats)
        
        db.session.commit()
        flash('Ticket stats saved.', 'success')
        return redirect(url_for('tickets'))
    
    return render_template('add_tickets.html', 
        agents=agents, today=datetime.now().strftime('%Y-%m-%d'))

@app.route('/tickets/sync/<int:agent_id>')
def sync_tickets(agent_id):
    """Sync tickets from Freshdesk for an agent."""
    success, result = sync_freshdesk_for_agent(agent_id)
    if success:
        flash(f'Synced: {result["handled"]} handled, {result["closed"]} closed', 'success')
    else:
        flash(f'Sync failed: {result}', 'danger')
    return redirect(url_for('tickets'))

@app.route('/tickets/fetch')
def fetch_tickets():
    """Fetch tickets from Freshdesk for all agents."""
    agents = Agent.query.filter_by(active=True).all()
    success_count = 0
    
    errors = []
    for agent in agents:
        success, result = sync_freshdesk_for_agent(agent.id)
        if success:
            success_count += 1
        else:
            errors.append(f'{agent.name}: {result}')

    if success_count > 0:
        flash(f'Synced tickets for {success_count} agent(s).', 'success')
    else:
        err_detail = '; '.join(errors[:3]) if errors else 'Unknown error'
        flash(f'Sync failed — {err_detail}', 'warning')
    
    return redirect(url_for('tickets'))

# ----- Weekly Reports -----

@app.route('/reports')
def reports():
    agent_id = get_active_agent_id()
    query = WeeklyReport.query
    if agent_id:
        query = query.filter_by(agent_id=agent_id)
    all_reports = query.order_by(WeeklyReport.week_ending.desc()).all()
    return render_template('reports.html', reports=all_reports)

@app.route('/reports/generate')
def generate_report():
    """Generate weekly reports for agents."""
    week_ending = get_week_ending()
    agent_id = get_active_agent_id()

    if agent_id:
        # Viewer: generate only for their agent
        generate_weekly_report_auto(agent_id, week_ending)
        flash(f'Generated your report for week ending {week_ending}', 'success')
    else:
        # Admin: generate for all agents
        agents = Agent.query.filter_by(active=True).all()
        for agent in agents:
            generate_weekly_report_auto(agent.id, week_ending)
        flash(f'Generated reports for week ending {week_ending}', 'success')
    return redirect(url_for('reports'))

@app.route('/reports/<int:id>')
def view_report(id):
    report = WeeklyReport.query.get_or_404(id)
    week_start = get_week_start(report.week_ending)
    week_start_dt = datetime.combine(week_start, datetime.min.time())
    week_end_dt = datetime.combine(report.week_ending, datetime.max.time())

    visits = SiteVisit.query.filter(
        SiteVisit.agent_id == report.agent_id,
        SiteVisit.date >= week_start,
        SiteVisit.date <= report.week_ending
    ).order_by(SiteVisit.date).all()

    ticket_stats = TicketStats.query.filter(
        TicketStats.agent_id == report.agent_id,
        TicketStats.date >= week_start,
        TicketStats.date <= report.week_ending
    ).all()

    logs = MonitoringLog.query.filter(
        MonitoringLog.timestamp >= week_start_dt,
        MonitoringLog.timestamp <= week_end_dt
    ).order_by(MonitoringLog.timestamp.desc()).all()

    calls = ServiceCall.query.filter(
        ServiceCall.agent_id == report.agent_id,
        ServiceCall.date >= week_start,
        ServiceCall.date <= report.week_ending
    ).order_by(ServiceCall.date.desc()).all()

    # Recalculate summary stats live so they're always current
    report.sites_visited = len(visits)
    monitoring_metrics = calculate_weekly_monitoring_metrics(week_start_dt, week_end_dt)
    report.sites_monitored = monitoring_metrics['sites_monitored']
    report.aps_offline_total = monitoring_metrics['aps_offline_total']
    report.critical_incidents = monitoring_metrics['critical_incidents']
    report.tickets_handled = sum(s.tickets_handled for s in ticket_stats)
    report.tickets_closed = sum(s.tickets_closed for s in ticket_stats)
    db.session.commit()

    holiday_map = get_public_holiday_map()
    week_holidays = []
    for day_str, name in holiday_map.items():
        day = _parse_iso_date(day_str)
        if day and week_start <= day <= report.week_ending:
            week_holidays.append({'date': day, 'name': name})
    week_holidays.sort(key=lambda x: x['date'])

    return render_template('view_report.html',
        report=report, visits=visits, ticket_stats=ticket_stats, logs=logs, calls=calls,
        week_holidays=week_holidays, is_pdf=False)


@app.route('/reports/<int:id>/export/pdf')
def export_report_pdf(id):
    """Open a one-page print-ready weekly report so users can Save as PDF manually."""
    report = WeeklyReport.query.get_or_404(id)
    week_start = get_week_start(report.week_ending)
    week_start_dt = datetime.combine(week_start, datetime.min.time())
    week_end_dt = datetime.combine(report.week_ending, datetime.max.time())

    visits = SiteVisit.query.filter(
        SiteVisit.agent_id == report.agent_id,
        SiteVisit.date >= week_start,
        SiteVisit.date <= report.week_ending
    ).order_by(SiteVisit.date).all()

    ticket_stats = TicketStats.query.filter(
        TicketStats.agent_id == report.agent_id,
        TicketStats.date >= week_start,
        TicketStats.date <= report.week_ending
    ).all()

    logs = MonitoringLog.query.filter(
        MonitoringLog.timestamp >= week_start_dt,
        MonitoringLog.timestamp <= week_end_dt
    ).order_by(MonitoringLog.timestamp.desc()).all()

    calls = ServiceCall.query.filter(
        ServiceCall.agent_id == report.agent_id,
        ServiceCall.date >= week_start,
        ServiceCall.date <= report.week_ending
    ).order_by(ServiceCall.date.desc()).all()

    # Keep displayed totals in sync before exporting.
    report.sites_visited = len(visits)
    monitoring_metrics = calculate_weekly_monitoring_metrics(week_start_dt, week_end_dt)
    report.sites_monitored = monitoring_metrics['sites_monitored']
    report.aps_offline_total = monitoring_metrics['aps_offline_total']
    report.critical_incidents = monitoring_metrics['critical_incidents']
    report.tickets_handled = sum(s.tickets_handled for s in ticket_stats)
    report.tickets_closed = sum(s.tickets_closed for s in ticket_stats)
    db.session.commit()

    # Keep the exported PDF concise so it fits on one page.
    pdf_visits = visits[:5]
    pdf_logs = logs[:10]
    pdf_calls = calls[:6]

    return render_template('report_pdf.html',
        report=report,
        visits=pdf_visits,
        logs=pdf_logs,
        calls=pdf_calls,
        total_visits=len(visits),
        total_logs=len(logs),
        total_calls=len(calls))

@app.route('/reports/<int:id>/submit', methods=['POST'])
def submit_report(id):
    report = WeeklyReport.query.get_or_404(id)
    tickets_handled_override = (request.form.get('tickets_handled_override') or '').strip()
    tickets_closed_override = (request.form.get('tickets_closed_override') or '').strip()
    if tickets_handled_override != '':
        report.tickets_handled = max(0, int(tickets_handled_override))
    if tickets_closed_override != '':
        report.tickets_closed = max(0, int(tickets_closed_override))
    report.achievements = request.form.get('achievements', '')
    report.challenges = request.form.get('challenges', '')
    report.notes = request.form.get('notes', '')
    report.submitted_at = datetime.utcnow()
    db.session.commit()
    flash('Report submitted.', 'success')
    return redirect(url_for('view_report', id=id))

# ----- Company Summary -----

@app.route('/summary')
@admin_required
def weekly_summary():
    week_ending = request.args.get('week')
    if week_ending:
        week_ending_date = datetime.strptime(week_ending, '%Y-%m-%d').date()
    else:
        week_ending_date = get_week_ending()
    
    reports = WeeklyReport.query.filter_by(week_ending=week_ending_date).all()
    agents = Agent.query.filter_by(active=True).all()
    
    week_start = get_week_start(week_ending_date)
    week_start_dt = datetime.combine(week_start, datetime.min.time())
    week_end_dt = datetime.combine(week_ending_date, datetime.max.time())
    monitoring_metrics = calculate_weekly_monitoring_metrics(week_start_dt, week_end_dt)

    totals = {
        'sites_visited': sum(r.sites_visited for r in reports),
        'sites_monitored': monitoring_metrics['sites_monitored'],
        'aps_offline': monitoring_metrics['aps_offline_total'],
        'critical_incidents': monitoring_metrics['critical_incidents'],
        'tickets_handled': sum(r.tickets_handled for r in reports),
        'tickets_closed': sum(r.tickets_closed for r in reports)
    }
    
    submitted_ids = [r.agent_id for r in reports]
    missing_agents = [a for a in agents if a.id not in submitted_ids]

    holiday_map = get_public_holiday_map()
    week_holidays = []
    for day_str, name in holiday_map.items():
        day = _parse_iso_date(day_str)
        if day and week_start <= day <= week_ending_date:
            week_holidays.append({'date': day, 'name': name})
    week_holidays.sort(key=lambda x: x['date'])
    
    return render_template('summary.html',
        reports=reports,
        totals=totals,
        week_ending=week_ending_date,
        missing_agents=missing_agents,
        week_holidays=week_holidays
    )

@app.route('/summary/export')
def export_summary():
    week_ending = request.args.get('week', get_week_ending().isoformat())
    week_ending_date = datetime.strptime(week_ending, '%Y-%m-%d').date()
    reports = WeeklyReport.query.filter_by(week_ending=week_ending_date).all()
    
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(['Agent', 'Sites Visited', 'Sites Monitored', 'APs Offline',
                    'Critical Incidents', 'Tickets Handled', 'Tickets Closed'])
    
    for r in reports:
        writer.writerow([r.agent.name, r.sites_visited, r.sites_monitored,
                        r.aps_offline_total, r.critical_incidents,
                        r.tickets_handled, r.tickets_closed])
    
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename=summary_{week_ending}.csv'}
    )


@app.route('/references')
def references():
    docs = ReferenceDocument.query.order_by(ReferenceDocument.uploaded_at.desc()).all()
    return render_template('references.html', docs=docs)


@app.route('/references/upload', methods=['POST'])
def upload_reference():
    _ensure_reference_docs_folder()

    uploaded = request.files.get('document')
    title = (request.form.get('title') or '').strip()
    source = (request.form.get('source') or 'manual').strip() or 'manual'
    notes = (request.form.get('notes') or '').strip()

    if not uploaded or not uploaded.filename:
        flash('Please select a file to upload.', 'warning')
        return redirect(url_for('references'))

    if not _is_allowed_reference_file(uploaded.filename):
        flash('File type not allowed. Use xlsx/xls/csv/pdf/txt/md/json.', 'warning')
        return redirect(url_for('references'))

    safe_original = secure_filename(uploaded.filename)
    timestamp = datetime.utcnow().strftime('%Y%m%d%H%M%S')
    stored_filename = f"{timestamp}_{safe_original}"
    save_path = os.path.join(app.config['REFERENCE_DOCS_FOLDER'], stored_filename)
    uploaded.save(save_path)

    doc = ReferenceDocument(
        title=title or os.path.splitext(safe_original)[0],
        source=source,
        notes=notes,
        original_filename=uploaded.filename,
        stored_filename=stored_filename
    )
    db.session.add(doc)
    db.session.commit()

    flash('Reference document uploaded.', 'success')
    return redirect(url_for('references'))


@app.route('/references/<int:id>/view')
def view_reference(id):
    doc = ReferenceDocument.query.get_or_404(id)
    _ensure_reference_docs_folder()
    preview = _load_reference_table_preview(doc)
    return render_template('reference_view.html', doc=doc, preview=preview)


@app.route('/references/<int:id>/download')
def download_reference(id):
    doc = ReferenceDocument.query.get_or_404(id)
    _ensure_reference_docs_folder()
    return send_from_directory(
        app.config['REFERENCE_DOCS_FOLDER'],
        doc.stored_filename,
        as_attachment=True,
        download_name=doc.original_filename
    )


@app.route('/references/<int:id>/delete', methods=['POST'])
def delete_reference(id):
    doc = ReferenceDocument.query.get_or_404(id)
    file_path = os.path.join(app.config['REFERENCE_DOCS_FOLDER'], doc.stored_filename)
    if os.path.exists(file_path):
        os.remove(file_path)
    db.session.delete(doc)
    db.session.commit()
    flash('Reference document deleted.', 'success')
    return redirect(url_for('references'))

# ----- Settings -----

@app.route('/settings', methods=['GET', 'POST'])
@admin_required
def settings():
    agents = Agent.query.all()
    
    if request.method == 'POST':
        section = request.form.get('section', '')
        if section == 'freshdesk':
            raw_domain = re.sub(r'^https?://', '', request.form.get('freshdesk_domain', '').strip().rstrip('/'))
            set_config('freshdesk_domain', raw_domain)
            set_config('freshdesk_api_key', request.form.get('freshdesk_api_key', '').strip())
            set_config('freshdesk_agent_email', request.form.get('freshdesk_agent_email', '').strip())
            set_config('freshdesk_self_agent_id', '')
        elif section == 'report':
            set_config('report_due_day', request.form.get('report_due_day', 'friday'))
            set_config('report_due_time', request.form.get('report_due_time', '14:00'))
        elif section == 'standby_rates':
            set_config('standby_rate_weekday', request.form.get('standby_rate_weekday', '350').strip())
            set_config('standby_rate_saturday', request.form.get('standby_rate_saturday', '500').strip())
            set_config('standby_rate_sunday', request.form.get('standby_rate_sunday', '700').strip())
            set_config('standby_rate_public_holiday', request.form.get('standby_rate_public_holiday', '700').strip())
            set_config('standby_public_holidays', request.form.get('standby_public_holidays', '').strip())
        elif section == 'branding':
            app_name = request.form.get('app_name', '').strip()
            company_name = request.form.get('company_name', '').strip()
            if app_name:
                set_config('app_name', app_name)
            if company_name:
                set_config('company_name', company_name)
            # Logo upload — stored as base64 data-URI in Config table
            # (avoids ephemeral filesystem on Render)
            logo_file = request.files.get('app_logo')
            if logo_file and logo_file.filename:
                allowed_mime = {'image/png', 'image/jpeg', 'image/jpg', 'image/svg+xml'}
                mime = logo_file.mimetype or ''
                if mime not in allowed_mime:
                    flash('Logo must be PNG, JPG, or SVG.', 'danger')
                    return redirect(url_for('settings'))
                import base64
                logo_bytes = logo_file.read()
                if len(logo_bytes) > 500_000:  # 500 KB limit
                    flash('Logo file is too large (max 500 KB).', 'danger')
                    return redirect(url_for('settings'))
                logo_b64 = base64.b64encode(logo_bytes).decode('utf-8')
                data_uri = f'data:{mime};base64,{logo_b64}'
                set_config('app_logo', data_uri)
            elif request.form.get('remove_logo'):
                set_config('app_logo', '')
        flash('Settings saved.', 'success')
        return redirect(url_for('settings'))
    
    config = {
        'freshdesk_domain': get_config('freshdesk_domain', ''),
        'freshdesk_api_key': get_config('freshdesk_api_key', ''),
        'freshdesk_agent_email': get_config('freshdesk_agent_email', ''),
        'report_due_day': get_config('report_due_day', 'friday'),
        'report_due_time': get_config('report_due_time', '14:00'),
        'standby_rate_weekday': get_config('standby_rate_weekday', '350'),
        'standby_rate_saturday': get_config('standby_rate_saturday', '500'),
        'standby_rate_sunday': get_config('standby_rate_sunday', '700'),
        'standby_rate_public_holiday': get_config('standby_rate_public_holiday', '700'),
        'standby_public_holidays': get_config('standby_public_holidays', ''),
        'public_holidays_json': get_config('public_holidays_json', json.dumps(DEFAULT_PUBLIC_HOLIDAYS, indent=2)),
        'leave_approval_email': get_config('leave_approval_email', 'luke@company.com'),
        'app_name': get_config('app_name', 'VOD Operations Portal'),
        'company_name': get_config('company_name', 'Vodacom'),
        'app_logo': get_config('app_logo', ''),
    }
    
    return render_template('settings.html', config=config, agents=agents)

@app.route('/settings/test-freshdesk')
@admin_required
def test_freshdesk():
    """Test Freshdesk connectivity and resolve the agent ID."""
    base = _freshdesk_base_url()
    api_key = get_config('freshdesk_api_key')
    if not base or not api_key:
        flash('Freshdesk domain and API key must be saved first.', 'warning')
        return redirect(url_for('settings'))
    # Force re-resolve so the test is always live.
    set_config('freshdesk_self_agent_id', '')
    agent_id, err = resolve_self_freshdesk_agent_id()
    if agent_id:
        flash(f'Freshdesk connection OK. Your Freshdesk agent ID: {agent_id}', 'success')
    else:
        flash(f'Freshdesk connection failed ({base}/api/v2/agents/me): {err}', 'danger')
    return redirect(url_for('settings'))


@app.route('/settings/save', methods=['POST'])
@admin_required
def save_settings():
    """Save settings (redirect target)."""
    section = request.form.get('section', '')
    if section == 'freshdesk':
        # Strip trailing slashes/protocols so URL is always clean.
        raw_domain = request.form.get('freshdesk_domain', '').strip().rstrip('/').lstrip('https://').lstrip('http://')
        set_config('freshdesk_domain', raw_domain)
        set_config('freshdesk_api_key', request.form.get('freshdesk_api_key', '').strip())
        set_config('freshdesk_agent_email', request.form.get('freshdesk_agent_email', '').strip())
        # Clear cached agent ID so it gets re-resolved with the new credentials.
        set_config('freshdesk_self_agent_id', '')
    elif section == 'report':
        set_config('report_due_day', request.form.get('report_due_day', 'friday'))
        set_config('report_due_time', request.form.get('report_due_time', '14:00'))
    elif section == 'standby_rates':
        try:
            weekday = float(request.form.get('standby_rate_weekday', '0').strip() or '0')
            saturday = float(request.form.get('standby_rate_saturday', '0').strip() or '0')
            sunday = float(request.form.get('standby_rate_sunday', '0').strip() or '0')
            public_holiday = float(request.form.get('standby_rate_public_holiday', '0').strip() or '0')
        except ValueError:
            flash('Invalid rate value — please enter numbers only.', 'warning')
            return redirect(url_for('settings'))
        set_config('standby_rate_weekday', str(weekday))
        set_config('standby_rate_saturday', str(saturday))
        set_config('standby_rate_sunday', str(sunday))
        set_config('standby_rate_public_holiday', str(public_holiday))
        set_config('standby_public_holidays', request.form.get('standby_public_holidays', '').strip())
    elif section == 'public_holidays':
        raw_json = request.form.get('public_holidays_json', '').strip()
        try:
            parsed = json.loads(raw_json or '{}')
            if not isinstance(parsed, dict):
                raise ValueError('Holiday JSON must be a date-to-name object')
            cleaned = {}
            for date_raw, holiday_name in parsed.items():
                day = _parse_iso_date(str(date_raw).strip())
                if not day:
                    raise ValueError(f'Invalid holiday date: {date_raw}')
                cleaned[day.isoformat()] = str(holiday_name or '').strip() or 'Public Holiday'
            set_config('public_holidays_json', json.dumps(cleaned, indent=2))
            set_config('leave_approval_email', request.form.get('leave_approval_email', 'luke@company.com').strip())
        except Exception as exc:
            flash(f'Holiday settings not saved: {exc}', 'warning')
            return redirect(url_for('settings'))
    flash('Settings saved.', 'success')
    return redirect(url_for('settings'))


@app.route('/leave')
def leave_requests():
    agent_name = get_active_agent_name()
    query = LeaveRequest.query
    if agent_name:
        query = query.filter_by(agent_name=agent_name)
    leaves = query.order_by(LeaveRequest.created_at.desc()).all()
    today = datetime.now().date()
    return render_template(
        'leave_requests.html',
        leaves=leaves,
        engineers=active_engineers_for_date(today),
        approval_email=get_config('leave_approval_email', 'luke@company.com'),
        today=today,
        holiday_map=get_public_holiday_map()
    )


@app.route('/leave/export.xlsx')
@admin_required
def export_leave_calendar_excel():
    """Export leave calendar as Excel for management."""
    leaves = LeaveRequest.query.order_by(LeaveRequest.start_date.asc(), LeaveRequest.agent_name.asc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Leave Calendar'

    ws.append([
        'Engineer', 'Start Date', 'End Date', 'Status',
        'Coverage Agent', 'Approver', 'Reason', 'Decision Notes',
        'Requested At', 'Decided At'
    ])

    for leave in leaves:
        ws.append([
            leave.agent_name,
            leave.start_date.isoformat() if leave.start_date else '',
            leave.end_date.isoformat() if leave.end_date else '',
            leave.status,
            leave.coverage_agent or '',
            leave.approver or '',
            leave.reason or '',
            leave.decision_notes or '',
            leave.created_at.strftime('%Y-%m-%d %H:%M') if leave.created_at else '',
            leave.decided_at.strftime('%Y-%m-%d %H:%M') if leave.decided_at else '',
        ])

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = str(cell.value or '')
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 48)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    month_tag = datetime.now().strftime('%Y-%m')
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=leave_calendar_{month_tag}.xlsx'}
    )


@app.route('/leave/add', methods=['POST'])
def add_leave_request():
    agent_name = (request.form.get('agent_name') or '').strip()
    reason = (request.form.get('reason') or '').strip()
    start_date = _parse_iso_date((request.form.get('start_date') or '').strip())
    end_date = _parse_iso_date((request.form.get('end_date') or '').strip())

    if not agent_name or agent_name not in get_standby_engineers():
        flash('Please select a valid engineer.', 'warning')
        return redirect(url_for('leave_requests'))
    if not start_date or not end_date or end_date < start_date:
        flash('Please provide a valid leave date range.', 'warning')
        return redirect(url_for('leave_requests'))

    leave = LeaveRequest(
        agent_name=agent_name,
        start_date=start_date,
        end_date=end_date,
        reason=reason,
        status='pending',
        approver='Luke'
    )
    db.session.add(leave)
    db.session.commit()

    email_sent, email_error = send_leave_submission_email(leave)
    if email_sent:
        flash('Leave request submitted and emailed to Luke for approval.', 'success')
    else:
        flash(f'Leave request submitted. Email not sent ({email_error}).', 'warning')
    return redirect(url_for('leave_requests'))


@app.route('/leave/<int:id>/approve', methods=['POST'])
def approve_leave_request(id):
    leave = LeaveRequest.query.get_or_404(id)
    coverage_agent = (request.form.get('coverage_agent') or '').strip()
    decision_notes = (request.form.get('decision_notes') or '').strip()

    leave.status = 'approved'
    leave.decided_at = datetime.utcnow()
    leave.decision_notes = decision_notes
    leave.coverage_agent = coverage_agent or None

    updates = 0
    day = leave.start_date
    while day <= leave.end_date:
        existing = DutyRoster.query.filter_by(date=day).first()
        if existing and existing.agent_name == leave.agent_name:
            chosen_coverage = coverage_agent
            if not chosen_coverage or not is_engineer_active_for_date(chosen_coverage, day):
                alternatives = [
                    name for name in active_engineers_for_date(day).keys()
                    if name != leave.agent_name
                ]
                chosen_coverage = alternatives[0] if alternatives else leave.agent_name

            existing.agent_name = chosen_coverage
            leave_note = f'Coverage for {leave.agent_name} leave ({leave.start_date} to {leave.end_date})'
            existing.notes = f'{(existing.notes or "").strip()} | {leave_note}'.strip(' |')
            updates += 1
        day += timedelta(days=1)

    db.session.commit()
    flash(f'Leave approved. Updated {updates} roster day(s) for coverage.', 'success')
    return redirect(url_for('leave_requests'))


@app.route('/leave/<int:id>/reject', methods=['POST'])
def reject_leave_request(id):
    leave = LeaveRequest.query.get_or_404(id)
    leave.status = 'rejected'
    leave.decided_at = datetime.utcnow()
    leave.decision_notes = (request.form.get('decision_notes') or '').strip()
    db.session.commit()
    flash('Leave request rejected.', 'info')
    return redirect(url_for('leave_requests'))

@app.route('/agents/<int:id>/set-current')
def set_current_agent(id):
    """Set an agent as the current user."""
    # Clear is_current from all agents
    Agent.query.update({Agent.active: Agent.active})  # No-op to trigger update
    
    # For now just redirect (is_current would need a new column)
    flash('Agent selection saved.', 'success')
    return redirect(url_for('settings'))


# ----- Admin User Management -----

@app.route('/admin/users')
@admin_required
def admin_users():
    users = User.query.order_by(User.created_at.desc()).all()
    all_agents = Agent.query.filter_by(active=True).order_by(Agent.name).all()
    return render_template('admin_users.html', users=users, agents=all_agents)


@app.route('/admin/users/add', methods=['POST'])
@admin_required
def admin_add_user():
    email = (request.form.get('email') or '').strip().lower()
    password = (request.form.get('password') or '').strip()
    role = request.form.get('role', 'viewer')
    agent_id = request.form.get('agent_id', type=int)

    if not email or not password:
        flash('Email and password are required.', 'warning')
        return redirect(url_for('admin_users'))

    if len(password) < 6:
        flash('Password must be at least 6 characters.', 'warning')
        return redirect(url_for('admin_users'))

    if User.query.filter_by(email=email).first():
        flash('A user with that email already exists.', 'warning')
        return redirect(url_for('admin_users'))

    # Only superadmins can create admin/superadmin users
    if role in ('admin', 'superadmin') and not current_user.is_superadmin:
        flash('Only superadmins can create admin accounts.', 'warning')
        return redirect(url_for('admin_users'))

    # Non-admin users must have an agent assigned
    if role not in ('admin', 'superadmin') and not agent_id:
        flash('Non-admin users must be linked to an agent.', 'warning')
        return redirect(url_for('admin_users'))

    if agent_id and not Agent.query.get(agent_id):
        flash('Selected agent does not exist.', 'warning')
        return redirect(url_for('admin_users'))

    user = User(
        email=email,
        role=role,
        agent_id=agent_id if agent_id else None,
        created_by=current_user.id
    )
    user.set_password(password)
    db.session.add(user)
    db.session.commit()
    flash(f'User {email} created as {role}.', 'success')
    return redirect(url_for('admin_users'))


@app.route('/admin/users/<int:id>/toggle', methods=['POST'])
@admin_required
def admin_toggle_user(id):
    user = User.query.get_or_404(id)
    if not current_user.can_manage(user):
        flash('You do not have permission to modify this user.', 'warning')
        return redirect(url_for('admin_users'))
    user.active = not user.active
    db.session.commit()
    status = 'activated' if user.active else 'deactivated'
    flash(f'User {user.email} {status}.', 'success')
    return redirect(url_for('admin_users'))


@app.route('/admin/users/<int:id>/delete', methods=['POST'])
@admin_required
def admin_delete_user(id):
    user = User.query.get_or_404(id)
    if not current_user.can_manage(user):
        flash('You do not have permission to delete this user.', 'warning')
        return redirect(url_for('admin_users'))
    email = user.email
    db.session.delete(user)
    db.session.commit()
    flash(f'User {email} deleted.', 'success')
    return redirect(url_for('admin_users'))


# ----- Duty Roster -----

def _calc_hours(start_str, end_str):
    """Return float hours between two HH:MM strings (handles overnight)."""
    try:
        fmt = '%H:%M'
        s = datetime.strptime(start_str, fmt)
        e = datetime.strptime(end_str, fmt)
        diff = (e - s).total_seconds()
        if diff < 0:
            diff += 86400  # overnight
        return round(diff / 3600, 2)
    except Exception:
        return 0.0


def _build_month_calendar(year, month):
    """Return list-of-weeks where each week is 7 date slots (Mon=0 ... Sun=6)."""
    import calendar
    cal = calendar.Calendar(firstweekday=0)
    weeks = []
    for week in cal.monthdatescalendar(year, month):
        weeks.append(week)
    return weeks


def _parse_public_holidays(raw):
    """Parse comma/newline-separated YYYY-MM-DD values into a date set."""
    holidays = set()
    if not raw:
        return holidays
    for token in re.split(r'[\n,;]+', raw):
        token = token.strip()
        if not token:
            continue
        try:
            holidays.add(datetime.strptime(token, '%Y-%m-%d').date())
        except ValueError:
            continue
    return holidays


def _parse_iso_date(raw):
    if not raw:
        return None
    try:
        return datetime.strptime(raw, '%Y-%m-%d').date()
    except ValueError:
        return None


def get_public_holiday_map():
    """Load holiday map from config JSON; fallback to default ZA holidays."""
    configured = (get_config('public_holidays_json') or '').strip()
    if not configured:
        return dict(DEFAULT_PUBLIC_HOLIDAYS)
    try:
        parsed = json.loads(configured)
        if isinstance(parsed, dict):
            cleaned = {}
            for k, v in parsed.items():
                d = _parse_iso_date(str(k).strip())
                if d:
                    cleaned[d.isoformat()] = str(v or '').strip() or 'Public Holiday'
            return cleaned or dict(DEFAULT_PUBLIC_HOLIDAYS)
    except Exception:
        pass
    return dict(DEFAULT_PUBLIC_HOLIDAYS)


def get_public_holiday_dates():
    holiday_map = get_public_holiday_map()
    holidays = {_parse_iso_date(day) for day in holiday_map.keys()}
    holidays = {d for d in holidays if d is not None}
    # Keep backward compatibility with standby-specific holiday setting.
    holidays.update(_parse_public_holidays(get_config('standby_public_holidays') or ''))
    return holidays


def is_engineer_active_for_date(agent_name, on_date):
    engineers = get_standby_engineers()
    info = engineers.get(agent_name)
    if not info:
        return False
    start_date = _parse_iso_date(info.get('start_date'))
    end_date = _parse_iso_date(info.get('end_date'))
    if start_date and on_date < start_date:
        return False
    if end_date and on_date > end_date:
        return False
    return True


def active_engineers_for_date(on_date):
    return {
        name: info for name, info in get_standby_engineers().items()
        if is_engineer_active_for_date(name, on_date)
    }


def active_engineers_for_month(year, month):
    import calendar
    first_day = datetime(year, month, 1).date()
    last_day = datetime(year, month, calendar.monthrange(year, month)[1]).date()
    result = {}
    for name, info in get_standby_engineers().items():
        start_date = _parse_iso_date(info.get('start_date')) or first_day
        end_date = _parse_iso_date(info.get('end_date')) or last_day
        if start_date <= last_day and end_date >= first_day:
            result[name] = info
    return result


def send_telegram_alert(message):
    """Send escalation message to Telegram when bot config is available."""
    bot_token = (get_config('telegram_bot_token') or '').strip()
    chat_id = (get_config('telegram_chat_id') or '').strip()
    if not bot_token or not chat_id:
        return False, 'Telegram bot token/chat id not configured'
    try:
        resp = requests.post(
            f'https://api.telegram.org/bot{bot_token}/sendMessage',
            json={'chat_id': chat_id, 'text': message},
            timeout=15
        )
        if resp.status_code == 200:
            return True, None
        return False, f'Telegram API error {resp.status_code}'
    except Exception as exc:
        return False, str(exc)


def send_leave_submission_email(leave_request):
    """Send leave request email to approver (Luke) via SMTP config."""
    smtp_host = (get_config('smtp_host') or '').strip()
    smtp_port = int((get_config('smtp_port') or '587').strip() or '587')
    smtp_user = (get_config('smtp_user') or '').strip()
    smtp_password = (get_config('smtp_password') or '').strip()
    smtp_from = (get_config('smtp_from') or smtp_user or 'noreply@localhost').strip()
    approval_to = (get_config('leave_approval_email') or 'luke@company.com').strip()

    if not smtp_host:
        return False, 'SMTP host not configured (email skipped)'

    msg = EmailMessage()
    msg['Subject'] = f'Leave Request: {leave_request.agent_name} ({leave_request.start_date} to {leave_request.end_date})'
    msg['From'] = smtp_from
    msg['To'] = approval_to
    msg.set_content(
        f'Agent: {leave_request.agent_name}\n'
        f'Start: {leave_request.start_date}\n'
        f'End: {leave_request.end_date}\n'
        f'Status: {leave_request.status}\n\n'
        f'Reason:\n{leave_request.reason or "(none)"}\n'
    )

    try:
        with smtplib.SMTP(smtp_host, smtp_port, timeout=20) as server:
            server.starttls()
            if smtp_user:
                server.login(smtp_user, smtp_password)
            server.send_message(msg)
        return True, None
    except Exception as exc:
        return False, str(exc)


def _send_system_email(to_address, subject, body):
    """Reusable helper: send a system-generated email via the configured SMTP settings."""
    smtp_host = (get_config('smtp_host') or '').strip()
    smtp_port = int((get_config('smtp_port') or '587').strip() or '587')
    smtp_user = (get_config('smtp_user') or '').strip()
    smtp_password = (get_config('smtp_password') or '').strip()
    smtp_from = (get_config('smtp_from') or smtp_user or 'noreply@localhost').strip()

    if not smtp_host:
        return False, 'SMTP host not configured'
    if not to_address:
        return False, 'No recipient address'

    msg = EmailMessage()
    msg['Subject'] = subject
    msg['From'] = smtp_from
    msg['To'] = to_address
    msg.set_content(body)

    try:
        with smtplib.SMTP(smtp_host, smtp_port, timeout=20) as server:
            server.starttls()
            if smtp_user:
                server.login(smtp_user, smtp_password)
            server.send_message(msg)
        return True, None
    except Exception as exc:
        return False, str(exc)


def _day_type_for_date(day, public_holidays):
    if day in public_holidays:
        return 'public_holiday'
    wd = day.weekday()
    if wd == 5:
        return 'saturday'
    if wd == 6:
        return 'sunday'
    return 'weekday'


def _shift_window_for_day_type(day_type):
    """Return fixed shift window and hours by day type."""
    if day_type == 'weekday':
        return {'start': '17:00', 'end': '23:00', 'hours': 6.0}
    return {'start': '08:00', 'end': '23:00', 'hours': 15.0}


@app.route('/roster')
def duty_roster():
    from datetime import date as date_cls
    today = date_cls.today()
    year = int(request.args.get('year', today.year))
    month = int(request.args.get('month', today.month))

    weeks = _build_month_calendar(year, month)

    # Load all roster entries for this month
    import calendar
    first_day = date_cls(year, month, 1)
    last_day = date_cls(year, month, calendar.monthrange(year, month)[1])
    entries = DutyRoster.query.filter(
        DutyRoster.date >= first_day,
        DutyRoster.date <= last_day
    ).all()
    roster_map = {e.date: e for e in entries}
    active_engineers = active_engineers_for_month(year, month)
    public_holidays = get_public_holiday_map()

    # Prev / next month navigation
    prev_month = (month - 2) % 12 + 1
    prev_year = year - 1 if month == 1 else year
    next_month = month % 12 + 1
    next_year = year + 1 if month == 12 else year

    import calendar as cal_mod
    month_name = cal_mod.month_name[month]

    return render_template('duty_roster.html',
        year=year, month=month, month_name=month_name,
        weeks=weeks, roster_map=roster_map,
        engineers=active_engineers,
        all_engineers=get_standby_engineers(),
        public_holidays=public_holidays,
        prev_year=prev_year, prev_month=prev_month,
        next_year=next_year, next_month=next_month,
        today=today,
        current_agent_name=get_active_agent_name()
    )


@app.route('/roster/add', methods=['GET', 'POST'])
def add_roster_entry():
    from datetime import date as date_cls
    if request.method == 'POST':
        date_str = request.form.get('date', '').strip()
        agent_name = request.form.get('agent_name', '').strip()
        notes = request.form.get('notes', '').strip()

        if not date_str or not agent_name:
            flash('Date and engineer are required.', 'warning')
            return redirect(request.url)
        if agent_name not in get_standby_engineers():
            flash('Unknown engineer name.', 'warning')
            return redirect(request.url)

        try:
            entry_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            flash('Invalid date format.', 'warning')
            return redirect(request.url)

        if not is_engineer_active_for_date(agent_name, entry_date):
            flash(f'{agent_name} is not active on {entry_date}.', 'warning')
            return redirect(request.url)

        existing = DutyRoster.query.filter_by(date=entry_date).first()
        if existing:
            existing.agent_name = agent_name
            existing.notes = notes
            flash(f'Roster entry for {entry_date} updated to {agent_name}.', 'success')
        else:
            entry = DutyRoster(date=entry_date, agent_name=agent_name, notes=notes)
            db.session.add(entry)
            flash(f'Roster entry added: {agent_name} on {entry_date}.', 'success')
        db.session.commit()
        return redirect(url_for('duty_roster', year=entry_date.year, month=entry_date.month))

    prefill_date = request.args.get('date', '')
    entry_date = _parse_iso_date(prefill_date) or datetime.now().date()
    return render_template('add_roster_entry.html',
        engineers=active_engineers_for_date(entry_date), prefill_date=prefill_date)


@app.route('/roster/bulk', methods=['GET', 'POST'])
def bulk_roster():
    """Assign an entire month's night roster in one go via a form."""
    from datetime import date as date_cls
    import calendar
    if request.method == 'POST':
        year = int(request.form.get('year', date_cls.today().year))
        month = int(request.form.get('month', date_cls.today().month))
        last = calendar.monthrange(year, month)[1]
        updated = 0
        for day in range(1, last + 1):
            key = f'day_{day}'
            agent_name = request.form.get(key, '').strip()
            if not agent_name or agent_name not in get_standby_engineers():
                continue
            d = date_cls(year, month, day)
            if not is_engineer_active_for_date(agent_name, d):
                continue
            existing = DutyRoster.query.filter_by(date=d).first()
            if existing:
                existing.agent_name = agent_name
            else:
                db.session.add(DutyRoster(date=d, agent_name=agent_name))
            updated += 1
        db.session.commit()
        flash(f'{updated} roster entries saved for {calendar.month_name[month]} {year}.', 'success')
        return redirect(url_for('duty_roster', year=year, month=month))

    today = date_cls.today()
    year = int(request.args.get('year', today.year))
    month = int(request.args.get('month', today.month))
    last = calendar.monthrange(year, month)[1]
    days = []
    for day in range(1, last + 1):
        d = date_cls(year, month, day)
        existing = DutyRoster.query.filter_by(date=d).first()
        days.append({'date': d, 'day_num': day, 'entry': existing})

    import calendar as cal_mod
    month_names = {i: cal_mod.month_name[i] for i in range(1, 13)}
    return render_template('bulk_roster.html',
        year=year, month=month,
        month_name=cal_mod.month_name[month],
        month_names=month_names,
        days=days,
        engineers=active_engineers_for_month(year, month)
    )


@app.route('/roster/<int:id>/delete', methods=['POST'])
def delete_roster_entry(id):
    entry = DutyRoster.query.get_or_404(id)
    date_ref = entry.date
    db.session.delete(entry)
    db.session.commit()
    flash('Roster entry removed.', 'success')
    return redirect(url_for('duty_roster', year=date_ref.year, month=date_ref.month))


@app.route('/roster/assign-bulk', methods=['POST'])
def roster_bulk_assign():
    """Assign multiple selected dates to one engineer in a single POST."""
    from datetime import date as date_cls
    agent_name = request.form.get('agent_name', '').strip()
    year = int(request.form.get('year', date_cls.today().year))
    month = int(request.form.get('month', date_cls.today().month))
    dates = request.form.getlist('dates')

    if not agent_name or agent_name not in get_standby_engineers():
        flash('Unknown engineer.', 'warning')
        return redirect(url_for('duty_roster', year=year, month=month))
    if not dates:
        flash('No dates selected.', 'warning')
        return redirect(url_for('duty_roster', year=year, month=month))

    updated = 0
    for date_str in dates:
        try:
            d = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            continue
        if not is_engineer_active_for_date(agent_name, d):
            continue
        existing = DutyRoster.query.filter_by(date=d).first()
        if existing:
            existing.agent_name = agent_name
        else:
            db.session.add(DutyRoster(date=d, agent_name=agent_name))
        updated += 1
    db.session.commit()
    flash(f'{updated} day{"s" if updated != 1 else ""} assigned to {agent_name}.', 'success')
    return redirect(url_for('duty_roster', year=year, month=month))


# ----- Standby Claims -----

@app.route('/standby-claims')
def standby_claims():
    from datetime import date as date_cls
    import calendar
    today = date_cls.today()
    year = int(request.args.get('year', today.year))
    month = int(request.args.get('month', today.month))
    # Default to current user's linked agent, or first active roster engineer
    _fallback = Agent.query.filter_by(active=True, roster_enabled=True).order_by(Agent.id).first()
    default_agent = get_active_agent_name() or (_fallback.name if _fallback else '')
    agent_filter = request.args.get('agent', default_agent)

    first_day = date_cls(year, month, 1)
    last_day = date_cls(year, month, calendar.monthrange(year, month)[1])

    # Pull roster days for this engineer this month
    roster_entries = []
    if agent_filter and agent_filter in get_standby_engineers():
        roster_entries = DutyRoster.query.filter(
            DutyRoster.date >= first_day,
            DutyRoster.date <= last_day,
            DutyRoster.agent_name == agent_filter
        ).order_by(DutyRoster.date.asc()).all()

    standby_days = len(roster_entries)

    # Build roster day rows only.
    roster_dates = [r.date for r in roster_entries]
    all_dates = sorted(roster_dates)
    day_rows = []

    rates = {
        'weekday': float(get_config('standby_rate_weekday') or 0),
        'saturday': float(get_config('standby_rate_saturday') or 0),
        'sunday': float(get_config('standby_rate_sunday') or 0),
        'public_holiday': float(get_config('standby_rate_public_holiday') or 0),
    }
    public_holidays = get_public_holiday_dates()
    day_type_counts = {'weekday': 0, 'saturday': 0, 'sunday': 0, 'public_holiday': 0}

    standby_pay = 0.0
    total_hours = 0.0
    for d in all_dates:
        day_type = _day_type_for_date(d, public_holidays)
        shift = _shift_window_for_day_type(day_type)
        day_rate = rates.get(day_type, 0)
        day_type_counts[day_type] += 1
        standby_pay += day_rate
        total_hours += shift['hours']
        day_rows.append({
            'date': d,
            'on_roster': True,
            'day_type': day_type,
            'shift_start': shift['start'],
            'shift_end': shift['end'],
            'hours': shift['hours'],
            'day_rate': day_rate,
        })

    total_pay = standby_pay
    rates_configured = any(v > 0 for v in rates.values())

    prev_month = (month - 2) % 12 + 1
    prev_year = year - 1 if month == 1 else year
    next_month = month % 12 + 1
    next_year = year + 1 if month == 12 else year

    import calendar as cal_mod
    month_name = cal_mod.month_name[month]
    return render_template('standby_claims.html',
        day_rows=day_rows,
        year=year, month=month, month_name=month_name,
        today_iso=today.strftime('%Y-%m-%d'),
        total_hours=total_hours,
        standby_days=standby_days,
        standby_pay=standby_pay, total_pay=total_pay,
        rates=rates, day_type_counts=day_type_counts,
        rates_configured=rates_configured,
        engineers=get_standby_engineers(),
        public_holiday_names=get_public_holiday_map(),
        agent_filter=agent_filter,
        prev_year=prev_year, prev_month=prev_month,
        next_year=next_year, next_month=next_month
    )


@app.route('/standby-claims/add', methods=['GET', 'POST'])
def add_standby_claim():
    from datetime import date as date_cls
    if request.method == 'POST':
        agent_name = request.form.get('agent_name', '').strip()
        date_str = request.form.get('date', '').strip()
        call_received = request.form.get('call_received', '').strip()
        start_time = request.form.get('start_time', '').strip()
        end_time = request.form.get('end_time', '').strip()
        description = request.form.get('description', '').strip()
        ticket_reference = request.form.get('ticket_reference', '').strip()

        if not agent_name or not date_str or not start_time or not end_time:
            flash('Engineer, date, start time and end time are required.', 'warning')
            return redirect(request.url)
        if agent_name not in get_standby_engineers():
            flash('Unknown engineer.', 'warning')
            return redirect(request.url)
        try:
            claim_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            flash('Invalid date.', 'warning')
            return redirect(request.url)

        hours = _calc_hours(start_time, end_time)
        claim = StandbyClaim(
            agent_name=agent_name,
            date=claim_date,
            call_received=call_received or None,
            start_time=start_time,
            end_time=end_time,
            hours=hours,
            description=description,
            ticket_reference=ticket_reference
        )
        db.session.add(claim)
        db.session.commit()
        flash(f'Standby claim added ({hours:.2f} hrs).', 'success')
        return redirect(url_for('standby_claims', year=claim_date.year, month=claim_date.month))

    prefill_date = request.args.get('date', '')
    prefill_agent = request.args.get('agent', '')
    return render_template('add_standby_claim.html',
        engineers=get_standby_engineers(),
        prefill_date=prefill_date,
        prefill_agent=prefill_agent
    )


@app.route('/standby-claims/<int:id>/edit', methods=['GET', 'POST'])
def edit_standby_claim(id):
    claim = StandbyClaim.query.get_or_404(id)
    if request.method == 'POST':
        claim.agent_name = request.form.get('agent_name', claim.agent_name).strip()
        date_str = request.form.get('date', '').strip()
        try:
            claim.date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            flash('Invalid date.', 'warning')
            return redirect(request.url)
        claim.call_received = request.form.get('call_received', '').strip() or None
        claim.start_time = request.form.get('start_time', claim.start_time).strip()
        claim.end_time = request.form.get('end_time', claim.end_time).strip()
        claim.hours = _calc_hours(claim.start_time, claim.end_time)
        claim.description = request.form.get('description', '').strip()
        claim.ticket_reference = request.form.get('ticket_reference', '').strip()
        db.session.commit()
        flash('Claim updated.', 'success')
        return redirect(url_for('standby_claims', year=claim.date.year, month=claim.date.month))
    return render_template('add_standby_claim.html',
        engineers=get_standby_engineers(), claim=claim,
        prefill_date=claim.date.strftime('%Y-%m-%d'),
        prefill_agent=claim.agent_name
    )


@app.route('/standby-claims/<int:id>/delete', methods=['POST'])
def delete_standby_claim(id):
    claim = StandbyClaim.query.get_or_404(id)
    year, month = claim.date.year, claim.date.month
    db.session.delete(claim)
    db.session.commit()
    flash('Claim entry deleted.', 'success')
    return redirect(url_for('standby_claims', year=year, month=month))


@app.route('/standby-claims/export')
def export_standby_claim():
    from datetime import date as date_cls
    import calendar as cal_mod
    today = date_cls.today()
    year = int(request.args.get('year', today.year))
    month = int(request.args.get('month', today.month))
    agent_name = request.args.get('agent', '')

    if not agent_name or agent_name not in get_standby_engineers():
        flash('Please select an engineer to export a claim for.', 'warning')
        return redirect(url_for('standby_claims', year=year, month=month))

    first_day = date_cls(year, month, 1)
    last_day = date_cls(year, month, cal_mod.monthrange(year, month)[1])
    month_name = cal_mod.month_name[month]

    roster_entries = DutyRoster.query.filter(
        DutyRoster.agent_name == agent_name,
        DutyRoster.date >= first_day,
        DutyRoster.date <= last_day
    ).order_by(DutyRoster.date.asc()).all()

    rates = {
        'weekday': float(get_config('standby_rate_weekday') or 0),
        'saturday': float(get_config('standby_rate_saturday') or 0),
        'sunday': float(get_config('standby_rate_sunday') or 0),
        'public_holiday': float(get_config('standby_rate_public_holiday') or 0),
    }
    public_holidays = get_public_holiday_dates()
    day_type_counts = {'weekday': 0, 'saturday': 0, 'sunday': 0, 'public_holiday': 0}
    standby_pay = 0.0
    total_hours = 0.0
    day_rows = []
    for entry in roster_entries:
        day_type = _day_type_for_date(entry.date, public_holidays)
        shift = _shift_window_for_day_type(day_type)
        rate = rates.get(day_type, 0)
        day_type_counts[day_type] += 1
        standby_pay += rate
        total_hours += shift['hours']
        day_rows.append({
            'date': entry.date,
            'day_type': day_type,
            'shift_start': shift['start'],
            'shift_end': shift['end'],
            'hours': shift['hours'],
            'rate': rate,
            'amount': rate,
        })

    total_pay = standby_pay
    roster_count = len(roster_entries)

    employee_signature_name = (request.args.get('employee_signature_name') or '').strip() or agent_name
    employee_signature_date = (request.args.get('employee_signature_date') or '').strip()
    manager_signature_name = (request.args.get('manager_signature_name') or '').strip()
    manager_signature_date = (request.args.get('manager_signature_date') or '').strip()

    def _display_date(raw_value, fallback=''):
        if not raw_value:
            return fallback
        try:
            return datetime.strptime(raw_value, '%Y-%m-%d').strftime('%d %B %Y')
        except ValueError:
            return raw_value

    employee_signature_date_display = _display_date(employee_signature_date, today.strftime('%d %B %Y'))
    manager_signature_date_display = _display_date(manager_signature_date, '____________________')

    return render_template('standby_claim_export.html',
        agent_name=agent_name,
        year=year, month=month, month_name=month_name,
        day_rows=day_rows,
        total_hours=total_hours,
        roster_count=roster_count,
        rates=rates,
        day_type_counts=day_type_counts,
        standby_pay=standby_pay,
        total_pay=total_pay,
        employee_signature_name=employee_signature_name,
        employee_signature_date_display=employee_signature_date_display,
        manager_signature_name=manager_signature_name,
        manager_signature_date_display=manager_signature_date_display,
        engineers=get_standby_engineers(),
        generated_on=today
    )


# ----- API Endpoints -----

@app.route('/api/health')
def api_health():
    """API endpoint for health status."""
    summary, status = get_site_health_summary()
    return jsonify({
        'summary': summary,
        'sites': [{
            'name': s['site'].name,
            'status': s['status'],
            'aps_offline': s['aps_offline'],
            'total_aps': s['total_aps']
        } for s in status]
    })

@app.route('/api/import/ruckus', methods=['POST'])
def api_import_ruckus():
    """API endpoint for automated Ruckus import (for scheduled tasks)."""
    content = request.get_data(as_text=True)
    content_type = request.content_type or ''
    
    try:
        if 'json' in content_type:
            data = parse_ruckus_json(content)
        else:
            data = parse_ruckus_csv(content)
        
        result = import_ruckus_data(data)
        return jsonify({'status': 'success', **result})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 400


# ============== Scheduled Job Endpoints ==============
# Protected by CRON_SECRET bearer token for external schedulers
# (cron-job.org, GitHub Actions, UptimeRobot, etc.)

def _verify_cron_token():
    """Verify the Bearer token matches the CRON_SECRET env var."""
    if not CRON_SECRET:
        return False, jsonify({'status': 'error', 'message': 'CRON_SECRET not configured on server'}), 500
    auth = request.headers.get('Authorization', '')
    token = auth.replace('Bearer ', '', 1) if auth.startswith('Bearer ') else ''
    if not token or token != CRON_SECRET:
        return False, jsonify({'status': 'error', 'message': 'Unauthorized'}), 401
    return True, None, None


@app.route('/api/cron/freshdesk-sync', methods=['POST'])
def cron_freshdesk_sync():
    """Scheduled endpoint: sync Freshdesk tickets for all active agents.

    Trigger externally with:
        curl -X POST https://your-app.onrender.com/api/cron/freshdesk-sync \\
             -H "Authorization: Bearer YOUR_CRON_SECRET"
    """
    ok, *err = _verify_cron_token()
    if not ok:
        return err[0], err[1]

    agents = Agent.query.filter_by(active=True).all()
    results = []
    for agent in agents:
        success, result = sync_freshdesk_for_agent(agent.id)
        results.append({
            'agent': agent.name,
            'success': success,
            'detail': result if success else str(result)
        })

    succeeded = sum(1 for r in results if r['success'])
    return jsonify({
        'status': 'success' if succeeded > 0 else 'failed',
        'synced': succeeded,
        'total_agents': len(agents),
        'results': results
    })


@app.route('/api/cron/ruckus-import', methods=['POST'])
def cron_ruckus_import():
    """Scheduled endpoint: import Ruckus data posted as JSON or CSV body.

    Trigger externally with:
        curl -X POST https://your-app.onrender.com/api/cron/ruckus-import \\
             -H "Authorization: Bearer YOUR_CRON_SECRET" \\
             -H "Content-Type: application/json" \\
             -d @ruckus_export.json
    """
    ok, *err = _verify_cron_token()
    if not ok:
        return err[0], err[1]

    content = request.get_data(as_text=True)
    if not content or not content.strip():
        return jsonify({'status': 'error', 'message': 'No data in request body'}), 400

    content_type = request.content_type or ''
    try:
        if 'json' in content_type:
            data = parse_ruckus_json(content)
        else:
            data = parse_ruckus_csv(content)

        result = import_ruckus_data(data)
        return jsonify({'status': 'success', **result})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 400


@app.route('/api/cron/health', methods=['GET'])
def cron_health():
    """Health check / keep-alive endpoint for external pingers.

    Configure an uptime monitor (e.g. UptimeRobot, cron-job.org) to hit
    this endpoint every 14 minutes to prevent Render free-tier sleep.
    """
    return jsonify({'status': 'ok', 'timestamp': datetime.utcnow().isoformat()})


@app.route('/api/cron/email-monday-reminder', methods=['POST'])
def cron_email_monday_reminder():
    """Scheduled endpoint: Monday 08:00 — weekly plan reminder to admin users."""
    ok, *err = _verify_cron_token()
    if not ok:
        return err[0], err[1]

    admins = User.query.filter(
        User.role.in_(['admin', 'superadmin']),
        User.active == True
    ).all()
    if not admins:
        return jsonify({'status': 'skipped', 'reason': 'no admin users found'})

    today = datetime.utcnow().strftime('%d %B %Y')
    sent, failed = 0, []
    for admin in admins:
        subject = f'[VOD Ops] Weekly Plan Reminder \u2014 {today}'
        body = (
            f'Good morning,\n\n'
            f'This is your Monday reminder to review the weekly operations plan.\n\n'
            f'Key actions:\n'
            f'  \u2022 Review open escalations and assign owners\n'
            f'  \u2022 Check engineer leave schedule and adjust coverage\n'
            f'  \u2022 Confirm weekend site health alerts are actioned\n'
            f'  \u2022 Review Ruckus import data for anomalies\n\n'
            f'Open the portal: {APP_BASE_URL}/\n\n'
            f'-- VOD Operations Portal'
        )
        ok2, err2 = _send_system_email(admin.email, subject, body)
        if ok2:
            sent += 1
        else:
            failed.append({'email': admin.email, 'error': err2})

    return jsonify({'status': 'success' if sent > 0 else 'failed', 'sent': sent, 'failed': failed})


@app.route('/api/cron/email-friday-reminder', methods=['POST'])
def cron_email_friday_reminder():
    """Scheduled endpoint: Friday 13:00 — report due reminder to all active users."""
    ok, *err = _verify_cron_token()
    if not ok:
        return err[0], err[1]

    recipients = list({u.email for u in User.query.filter(User.active == True).all() if u.email})
    if not recipients:
        return jsonify({'status': 'skipped', 'reason': 'no recipients found'})

    today = datetime.utcnow().strftime('%d %B %Y')
    sent, failed = 0, []
    for email in recipients:
        subject = f'[VOD Ops] Weekly Report Due Today \u2014 {today}'
        body = (
            f'Reminder: the weekly report is due by end of business today ({today}).\n\n'
            f'Please ensure:\n'
            f'  \u2022 All service calls for the week are logged\n'
            f'  \u2022 Ticket statistics are up to date\n'
            f'  \u2022 Site visits are captured\n'
            f'  \u2022 Monitoring log is complete\n'
            f'  \u2022 Open escalations have status notes\n\n'
            f'Submit here: {APP_BASE_URL}/reports\n\n'
            f'-- VOD Operations Portal'
        )
        ok2, err2 = _send_system_email(email, subject, body)
        if ok2:
            sent += 1
        else:
            failed.append({'email': email, 'error': err2})

    return jsonify({'status': 'success' if sent > 0 else 'failed', 'sent': sent, 'failed': failed})


@app.route('/api/cron/email-monthly-summary', methods=['POST'])
def cron_email_monthly_summary():
    """Scheduled endpoint: 1st of month — ops summary to all active users."""
    ok, *err = _verify_cron_token()
    if not ok:
        return err[0], err[1]

    now = datetime.utcnow()
    first_this_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    first_last_month = (first_this_month - timedelta(days=1)).replace(day=1)
    last_month_label = first_last_month.strftime('%B %Y')

    tickets_total = db.session.query(
        db.func.sum(TicketStats.tickets_handled)
    ).filter(
        TicketStats.date >= first_last_month,
        TicketStats.date < first_this_month
    ).scalar() or 0

    service_calls_total = ServiceCall.query.filter(
        ServiceCall.date >= first_last_month,
        ServiceCall.date < first_this_month
    ).count()

    site_visits_total = SiteVisit.query.filter(
        SiteVisit.date >= first_last_month,
        SiteVisit.date < first_this_month
    ).count()

    aps_offline = Site.query.filter(Site.aps_offline > 0, Site.active == True).count()
    total_sites = Site.query.filter_by(active=True).count()

    recipients = list({u.email for u in User.query.filter(User.active == True).all() if u.email})
    if not recipients:
        return jsonify({'status': 'skipped', 'reason': 'no active users found'})

    sent, failed = 0, []
    for email in recipients:
        subject = f'[VOD Ops] Monthly Summary \u2014 {last_month_label}'
        body = (
            f'VOD Support \u2014 Monthly Operations Summary: {last_month_label}\n'
            f'{"=" * 50}\n\n'
            f'Tickets Handled:       {tickets_total}\n'
            f'Service Calls:         {service_calls_total}\n'
            f'Site Visits:           {site_visits_total}\n'
            f'Sites with AP Issues:  {aps_offline} / {total_sites}\n\n'
            f'Full detail: {APP_BASE_URL}/reports\n\n'
            f'-- VOD Operations Portal'
        )
        ok2, err2 = _send_system_email(email, subject, body)
        if ok2:
            sent += 1
        else:
            failed.append({'email': email, 'error': err2})

    return jsonify({'status': 'success' if sent > 0 else 'failed', 'sent': sent, 'failed': failed})


# ============== Wallboard Mode ==============

@app.route('/wallboard')
@login_required
def wallboard():
    """Office TV wallboard \u2014 full-screen live operations display (auto-refreshes every 60s)."""
    try:
        now = datetime.utcnow()
        today = now.date()

        total_sites = Site.query.filter_by(active=True).count() or 0
        sites_healthy = Site.query.filter(Site.active == True, Site.aps_offline == 0).count() or 0
        total_aps = int(db.session.query(db.func.sum(Site.total_aps)).filter(Site.active == True).scalar() or 0)
        total_offline = int(db.session.query(db.func.sum(Site.aps_offline)).filter(Site.active == True).scalar() or 0)
        health_pct = round((sites_healthy / total_sites * 100) if total_sites > 0 else 0)

        total_engineers = Agent.query.filter_by(active=True).count() or 0
        approved_today = LeaveRequest.query.filter(
            LeaveRequest.status == 'approved',
            LeaveRequest.start_date <= today,
            LeaveRequest.end_date >= today
        ).count() or 0
        engineers_on_duty = max(total_engineers - approved_today, 0)

        week_start = today - timedelta(days=today.weekday())
        escalated_this_week = int(db.session.query(
            db.func.sum(TicketStats.tickets_escalated)
        ).filter(
            TicketStats.date >= week_start,
            TicketStats.date <= today
        ).scalar() or 0)

        tickets_today = int(db.session.query(
            db.func.sum(TicketStats.tickets_handled)
        ).filter(
            TicketStats.date == today
        ).scalar() or 0)

        return render_template(
            'wallboard.html',
            health_pct=health_pct,
            sites_healthy=sites_healthy,
            total_sites=total_sites,
            total_aps=total_aps,
            total_offline=total_offline,
            engineers_on_duty=engineers_on_duty,
            total_engineers=total_engineers,
            escalated_this_week=escalated_this_week,
            tickets_today=tickets_today,
            now=now,
            refresh_seconds=60,
            wallboard_error=None,
        )
    except Exception as exc:
        app.logger.exception('Wallboard render error')
        return render_template(
            'wallboard.html',
            health_pct=0, sites_healthy=0, total_sites=0,
            total_aps=0, total_offline=0,
            engineers_on_duty=0, total_engineers=0,
            escalated_this_week=0, tickets_today=0,
            now=datetime.utcnow(), refresh_seconds=60,
            wallboard_error=str(exc),
        )

# ============== Initialize ==============

def _is_postgres():
    """Check if the current database is PostgreSQL (vs SQLite)."""
    return app.config['SQLALCHEMY_DATABASE_URI'].startswith('postgresql')

def _run_sqlite_migrations(conn):
    """Run SQLite-specific PRAGMA-based schema upgrades for local dev."""
    from sqlalchemy import text as _text
    try:
        sv_cols = [row[1] for row in conn.execute(_text('PRAGMA table_info(site_visit)'))]
        if 'location' not in sv_cols:
            conn.execute(_text('ALTER TABLE site_visit ADD COLUMN location VARCHAR(200)'))
            conn.commit()
    except Exception:
        conn.commit()
    try:
        ts_cols = [row[1] for row in conn.execute(_text('PRAGMA table_info(ticket_stats)'))]
        if 'tickets_pending' not in ts_cols:
            conn.execute(_text('ALTER TABLE ticket_stats ADD COLUMN tickets_pending INTEGER'))
            conn.commit()
    except Exception:
        conn.commit()
    try:
        u_cols = [row[1] for row in conn.execute(_text('PRAGMA table_info(app_user)'))]
        if 'agent_id' not in u_cols:
            conn.execute(_text('ALTER TABLE app_user ADD COLUMN agent_id INTEGER REFERENCES agent(id)'))
            conn.commit()
        if 'created_by' not in u_cols:
            conn.execute(_text('ALTER TABLE app_user ADD COLUMN created_by INTEGER'))
            conn.commit()
        if 'last_login' not in u_cols:
            conn.execute(_text('ALTER TABLE app_user ADD COLUMN last_login TIMESTAMP'))
            conn.commit()
    except Exception:
        conn.commit()
    # Standby / roster columns on agent table
    try:
        a_cols = [row[1] for row in conn.execute(_text('PRAGMA table_info(agent)'))]
        for _col, _type in [
            ('roster_enabled',         'BOOLEAN DEFAULT 1'),
            ('standby_color',          'VARCHAR(20)'),
            ('standby_text_color',     'VARCHAR(20)'),
            ('standby_label',          'VARCHAR(20)'),
            ('standby_start_date',     'DATE'),
            ('standby_end_date',       'DATE'),
            ('standby_responsibilities', 'VARCHAR(300)'),
        ]:
            if _col not in a_cols:
                conn.execute(_text(f'ALTER TABLE agent ADD COLUMN {_col} {_type}'))
                conn.commit()
    except Exception:
        conn.commit()

def _run_postgres_migrations(conn):
    """Run Postgres-specific schema upgrades using information_schema."""
    from sqlalchemy import text as _text
    def _col_exists(table, column):
        result = conn.execute(_text(
            "SELECT 1 FROM information_schema.columns "
            "WHERE table_name = :tbl AND column_name = :col"
        ), {'tbl': table, 'col': column})
        return result.fetchone() is not None

    if not _col_exists('site_visit', 'location'):
        conn.execute(_text('ALTER TABLE site_visit ADD COLUMN location VARCHAR(200)'))
        conn.commit()
    if not _col_exists('ticket_stats', 'tickets_pending'):
        conn.execute(_text('ALTER TABLE ticket_stats ADD COLUMN tickets_pending INTEGER'))
        conn.commit()
    if not _col_exists('app_user', 'agent_id'):
        conn.execute(_text('ALTER TABLE app_user ADD COLUMN agent_id INTEGER REFERENCES agent(id)'))
        conn.commit()
    if not _col_exists('app_user', 'created_by'):
        conn.execute(_text('ALTER TABLE app_user ADD COLUMN created_by INTEGER'))
        conn.commit()
    if not _col_exists('app_user', 'last_login'):
        conn.execute(_text('ALTER TABLE app_user ADD COLUMN last_login TIMESTAMP'))
        conn.commit()
    # Standby / roster columns on agent table
    for _col, _type in [
        ('roster_enabled',         'BOOLEAN DEFAULT TRUE'),
        ('standby_color',          'VARCHAR(20)'),
        ('standby_text_color',     'VARCHAR(20)'),
        ('standby_label',          'VARCHAR(20)'),
        ('standby_start_date',     'DATE'),
        ('standby_end_date',       'DATE'),
        ('standby_responsibilities', 'VARCHAR(300)'),
    ]:
        if not _col_exists('agent', _col):
            conn.execute(_text(f'ALTER TABLE agent ADD COLUMN {_col} {_type}'))
            conn.commit()

def init_db():
    with app.app_context():
        _ensure_reference_docs_folder()
        db.create_all()

        # Run schema migrations (column additions for existing databases)
        with db.engine.connect() as _conn:
            if _is_postgres():
                _run_postgres_migrations(_conn)
            else:
                _run_sqlite_migrations(_conn)
        
        # Create default agent
        if Agent.query.count() == 0:
            agent = Agent(name='Chris', email='chris@company.com')
            db.session.add(agent)
            db.session.commit()
            print("  Created default agent 'Chris'")
        
        # Seed assigned sites if empty
        if Site.query.count() == 0:
            print("  Seeding assigned sites...")
            for region, sites in ASSIGNED_SITES.items():
                for site_data in sites:
                    site = Site(
                        name=site_data['name'],
                        region=region,
                        ruckus_zone_name=', '.join(site_data['ruckus_zones'][:2]),
                        active=True
                    )
                    db.session.add(site)
            db.session.commit()
            print(f"  Added {Site.query.count()} assigned sites")

        # ── Seed users from SEED_USERS env var (runs on every startup; safe to leave set) ──
        # Format: email:password:role,email2:password2:role2
        # Roles:  superadmin | admin | viewer
        # Example:
        #   SEED_USERS=chris@vod.co.za:MyPass1:superadmin,luke@vod.co.za:MyPass2:admin,mgmt@vod.co.za:MyPass3:viewer
        _seed_spec = os.environ.get('SEED_USERS', '').strip()
        if _seed_spec:
            for _entry in _seed_spec.split(','):
                _entry = _entry.strip()
                if not _entry:
                    continue
                _parts = _entry.split(':')
                if len(_parts) < 3:
                    print(f"  SEED_USERS: skipping malformed entry (need email:password:role): {_entry}")
                    continue
                _email, _pwd, _role = _parts[0].strip(), _parts[1].strip(), _parts[2].strip()
                if _role not in ('superadmin', 'admin', 'viewer'):
                    print(f"  SEED_USERS: skipping {_email} — invalid role '{_role}'")
                    continue
                _existing = User.query.filter_by(email=_email).first()
                if _existing:
                    _existing.set_password(_pwd)
                    _existing.role = _role
                    _existing.active = True
                    db.session.commit()
                    print(f"  SEED_USERS: updated  {_email}  ({_role})")
                else:
                    _u = User(email=_email, role=_role)
                    _u.set_password(_pwd)
                    db.session.add(_u)
                    db.session.commit()
                    print(f"  SEED_USERS: created  {_email}  ({_role})")

        # Create initial superadmin only if NO users exist at all and SEED_USERS is not set.
        # This runs once on a brand-new database.
        elif User.query.count() == 0:
            admin = User(email='admin@vodacom.co.za', role='superadmin')
            admin.set_password('ChangeMe@2026!')
            db.session.add(admin)
            db.session.commit()
            print("  Created initial superadmin: admin@vodacom.co.za  (change password immediately)")


# ============== CLI Commands ==============
import click


@app.cli.command('create-user')
@click.argument('email')
@click.argument('role', default='viewer')
@click.option('--password', prompt=True, hide_input=True, confirmation_prompt=True)
def create_user_cmd(email, role, password):
    """Create or update an application user.

    ROLE: superadmin | admin | viewer  (default: viewer)

    Examples:

        flask create-user chris@vod.co.za superadmin

        flask create-user luke@vod.co.za admin

        flask create-user mgmt@vod.co.za viewer
    """
    if role not in ('superadmin', 'admin', 'viewer'):
        click.echo(f'Error: role must be superadmin, admin, or viewer. Got: {role}', err=True)
        raise SystemExit(1)
    existing = User.query.filter_by(email=email).first()
    if existing:
        existing.set_password(password)
        existing.role = role
        existing.active = True
        db.session.commit()
        click.echo(f'Updated:  {email}  ({role})')
    else:
        user = User(email=email, role=role)
        user.set_password(password)
        db.session.add(user)
        db.session.commit()
        click.echo(f'Created:  {email}  ({role})')


@app.cli.command('seed-users')
def seed_users_cmd():
    """Seed production users from the SEED_USERS environment variable.

    Format (comma-separated triples):

        SEED_USERS=email:password:role,email2:password2:role2

    Roles: superadmin | admin | viewer

    Example:

        SEED_USERS=chris@vod.co.za:SecurePass1:superadmin,luke@vod.co.za:SecurePass2:admin \\\
            flask seed-users
    """
    spec = os.environ.get('SEED_USERS', '').strip()
    if not spec:
        click.echo('SEED_USERS env var is not set \u2014 nothing to seed.', err=True)
        return

    created, updated = 0, 0
    for entry in spec.split(','):
        entry = entry.strip()
        if not entry:
            continue
        parts = entry.split(':')
        if len(parts) < 3:
            click.echo(f'Skipping malformed entry (need email:password:role): {entry}', err=True)
            continue
        email, password, role = parts[0].strip(), parts[1].strip(), parts[2].strip()
        if role not in ('superadmin', 'admin', 'viewer'):
            click.echo(f'Skipping {email}: invalid role \'{role}\'', err=True)
            continue
        existing = User.query.filter_by(email=email).first()
        if existing:
            existing.set_password(password)
            existing.role = role
            existing.active = True
            db.session.commit()
            click.echo(f'Updated:  {email}  ({role})')
            updated += 1
        else:
            user = User(email=email, role=role)
            user.set_password(password)
            db.session.add(user)
            db.session.commit()
            click.echo(f'Created:  {email}  ({role})')
            created += 1

    click.echo(f'\nDone \u2014 Created: {created}, Updated: {updated}')


# Initialize DB on import (required for gunicorn)
init_db()

if __name__ == '__main__':
    print("\n" + "="*60)
    print("  VOD Operations Portal v2.0")
    print("="*60)
    print("\n  Access: http://localhost:5000")
    print("  Press Ctrl+C to stop")
    print("="*60 + "\n")

    debug_mode = os.environ.get('FLASK_DEBUG', 'false').lower() == 'true'
    app.run(host='0.0.0.0', port=5000, debug=debug_mode)
